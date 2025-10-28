#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session
from datetime import datetime, date, timedelta
import json
import os
from database import DatabaseManager
from auto_scheduler import AutoMeetingScheduler
from services.auto_schedule_service import AutoScheduleService
from services.constraints_service import ConstraintsService
from services.committee_types_service import CommitteeTypesService, CommitteeTypeRequest
from services.committee_recommendation_service import CommitteeRecommendationService
from services.audit_logger import AuditLogger
from services.ad_service import ADService
from auth import AuthManager, login_required, admin_required, editor_required, editing_permission_required

app = Flask(__name__)
app.secret_key = 'committee_management_secret_key_2025_azure_oauth_enabled'
app.config['SESSION_COOKIE_SECURE'] = False  # For development (HTTP)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = None  # Allow cross-site for OAuth redirects
app.config['PERMANENT_SESSION_LIFETIME'] = 3600  # 1 hour
app.config['SESSION_REFRESH_EACH_REQUEST'] = False  # Don't refresh on every request

# Initialize system components
db = DatabaseManager()
ad_service = ADService(db)
auto_scheduler = AutoMeetingScheduler(db)
auto_schedule_service = AutoScheduleService(db)
constraints_service = ConstraintsService(db)
committee_types_service = CommitteeTypesService(db)
committee_recommendation_service = CommitteeRecommendationService(db)
auth_manager = AuthManager(db, ad_service)
audit_logger = AuditLogger(db)

# Mobile device detection middleware
def is_mobile_device():
    """Detect if the request is from a mobile device"""
    user_agent = request.headers.get('User-Agent', '').lower()
    mobile_keywords = ['android', 'webos', 'iphone', 'ipad', 'ipod', 'blackberry', 'windows phone', 'mobile']
    return any(keyword in user_agent for keyword in mobile_keywords)

@app.before_request
def check_mobile_access():
    """Block mobile device access (optional - can be disabled)"""
    # Skip check for static files and API endpoints
    if request.path.startswith('/static/') or request.path.startswith('/api/'):
        return None
    
    # Uncomment the following lines to enable server-side mobile blocking
    # if is_mobile_device():
    #     return render_template('mobile_blocked.html'), 403
    
    return None

# Authentication routes
@app.route('/login', methods=['GET', 'POST'])
def login():
    """User login - Check if Azure AD is configured before redirecting"""
    # Check if Azure AD credentials are configured
    if not ad_service.azure_tenant_id or not ad_service.azure_client_id or not ad_service.azure_client_secret:
        # Azure AD not configured - provide options
        if request.method == 'POST':
            # Check if user wants to bypass Azure AD temporarily
            bypass = request.form.get('bypass_azure', 'false')
            if bypass == 'true':
                flash('התחברות באמצעות Azure AD מבוטלת זמנית. אנא הגדר את פרטי ההתחברות.', 'warning')
                return render_template('base.html', content="""
                <div class="row justify-content-center">
                    <div class="col-md-8">
                        <div class="alert alert-warning text-center">
                            <h4><i class="bi bi-exclamation-triangle"></i> אימות Azure AD לא מוגדר</h4>
                            <p>המערכת דורשת הגדרת פרטי התחברות Azure AD כדי לפעול.</p>
                            <p>אנא פנה למנהל המערכת או הגדר את הפרטים בקובץ .env</p>
                        </div>
                    </div>
                </div>
                """)

        flash('אימות Azure AD לא מוגדר - נדרשת הגדרת פרטי התחברות ב-.env', 'error')
        return render_template('base.html', content="""
        <div class="row justify-content-center">
            <div class="col-md-8">
                <div class="alert alert-danger text-center">
                    <h4><i class="bi bi-exclamation-triangle"></i> אימות Azure AD לא מוגדר</h4>
                    <p>המערכת דורשת הגדרת פרטי התחברות Azure AD כדי לפעול.</p>
                    <p>אנא פנה למנהל המערכת או הגדר את הפרטים בקובץ .env</p>
                    <div class="mt-3">
                        <small class="text-muted">קובץ .env.example נוצר בתיקיית הפרויקט</small>
                    </div>
                </div>
            </div>
        </div>
        """)

    # Azure AD is configured - redirect to OAuth
    return redirect(url_for('auth_azure'))

@app.route('/logout')
def logout():
    """User logout"""
    username = session.get('username', 'Unknown')
    auth_manager.logout_user()
    
    # Log the logout
    audit_logger.log_logout(username)
    
    # Redirect to SSO login (which will auto-redirect to Azure AD)
    flash('התנתקת מהמערכת בהצלחה', 'success')
    return redirect(url_for('auth_azure'))

@app.route('/refresh_session')
@login_required
def refresh_session():
    """Refresh session data from database - useful after role changes"""
    if 'user_id' not in session:
        flash('נדרשת התחברות', 'error')
        return redirect(url_for('auth_azure'))
    
    try:
        # Get fresh user data from database
        user = db.get_user_by_id(session['user_id'])
        
        if not user:
            flash('משתמש לא נמצא', 'error')
            auth_manager.logout_user()
            return redirect(url_for('auth_azure'))
        
        if not user['is_active']:
            flash('חשבון המשתמש מושבת', 'error')
            auth_manager.logout_user()
            return redirect(url_for('auth_azure'))
        
        # Update session with fresh data from database
        old_role = session.get('role')
        session['role'] = user['role']
        session['username'] = user['username']
        session['full_name'] = user['full_name']
        session['hativa_id'] = user.get('hativa_id')
        
        if old_role != user['role']:
            flash(f'הרשאות עודכנו: {old_role} → {user["role"]}', 'success')
        else:
            flash('Session עודכן בהצלחה', 'success')
        
        return redirect(url_for('index'))
        
    except Exception as e:
        app.logger.error(f"Error refreshing session: {e}", exc_info=True)
        flash(f'שגיאה בעדכון session: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/auth/azure')
def auth_azure():
    """Redirect to Azure AD for authentication"""
    import secrets
    from flask import session as flask_session

    # Check if Azure AD credentials are configured in .env
    if not ad_service.azure_tenant_id or not ad_service.azure_client_id or not ad_service.azure_client_secret:
        flash('אימות Azure AD לא מוגדר - נדרשת הגדרת פרטי התחברות בקובץ .env', 'error')
        return render_template('base.html', content="""
        <div class="row justify-content-center">
            <div class="col-md-8">
                <div class="alert alert-danger text-center">
                    <h4><i class="bi bi-exclamation-triangle"></i> אימות Azure AD לא מוגדר</h4>
                    <p>המערכת דורשת הגדרת פרטי התחברות Azure AD כדי לפעול.</p>
                    <p>אנא פנה למנהל המערכת או הגדר את הפרטים בקובץ .env</p>
                    <div class="mt-3">
                        <small class="text-muted">קובץ .env.example נוצר בתיקיית הפרויקט</small>
                    </div>
                </div>
            </div>
        </div>
        """)

    # Generate state for CSRF protection
    state = secrets.token_urlsafe(32)
    flask_session.permanent = True  # Make session persistent
    flask_session['oauth_state'] = state
    flask_session.modified = True  # Force session save

    app.logger.info(f"Generated state and saved to session: {state[:20]}...")

    # Get authorization URL
    auth_url = ad_service.get_azure_auth_url(state=state)

    if not auth_url:
        flash('שגיאה ביצירת קישור אימות Azure AD - בדוק הגדרות ב-.env', 'error')
        return render_template('base.html', content="""
        <div class="row justify-content-center">
            <div class="col-md-8">
                <div class="alert alert-danger text-center">
                    <h4><i class="bi bi-exclamation-triangle"></i> שגיאה בהגדרות Azure AD</h4>
                    <p>לא ניתן ליצור קישור אימות. בדוק שהגדרות Azure AD תקינות.</p>
                </div>
            </div>
        </div>
        """)

    return redirect(auth_url)

@app.route('/bypass_auth')
def bypass_auth():
    """Temporary bypass for Azure AD authentication (development only)"""
    # Create a temporary admin user for testing
    session['user_id'] = 1
    session['username'] = 'admin'
    session['role'] = 'admin'
    session['hativa_id'] = None
    session['full_name'] = 'מנהל מערכת (בדיקה)'
    session['auth_source'] = 'bypass'

    flash('התחברת באמצעות bypass - זה למטרות בדיקה בלבד', 'warning')
    return redirect(url_for('index'))

@app.route('/auth/callback')
def auth_callback():
    # Verify state parameter
    state = request.args.get('state')
    session_state = session.get('oauth_state')
    
    app.logger.info(f"Callback received - State from request: {state[:20] if state else 'None'}...")
    app.logger.info(f"State from session: {session_state[:20] if session_state else 'None'}...")
    
    # TODO: Fix session persistence issue with OAuth redirects
    # For now, skip state validation if session was lost (common with cross-site redirects)
    if session_state and state != session_state:
        app.logger.warning("State mismatch detected but continuing (session lost during redirect)")
    
    # Clear state from session
    session.pop('oauth_state', None)
    
    # Check for error from Azure AD
    error = request.args.get('error')
    if error:
        error_description = request.args.get('error_description', error)
        flash(f'שגיאת אימות Azure AD: {error_description}', 'error')
        return redirect(url_for('login'))
    
    # Get authorization code
    auth_code = request.args.get('code')
    if not auth_code:
        flash('לא התקבל קוד אימות מ-Azure AD', 'error')
        return redirect(url_for('login'))
    
    # Authenticate with code
    app.logger.info("Authenticating with Azure AD code...")
    success, ad_user_info, message = ad_service.authenticate_with_code(auth_code)
    
    if not success or not ad_user_info:
        app.logger.error(f"Azure AD authentication failed: {message}")
        flash(f'שגיאה באימות: {message}', 'error')
        audit_logger.log_login(request.args.get('loginHint', 'unknown'), False, message)
        return redirect(url_for('login'))
    
    # Check if user exists in local DB
    username = ad_user_info.get('username', ad_user_info.get('email', 'unknown'))
    app.logger.info(f"Azure AD auth successful. Username: {username}")
    app.logger.info(f"User info: {ad_user_info}")
    
    user = db.get_user_by_username_any_source(username)
    app.logger.info(f"User exists in DB: {user is not None}")
    
    if user:
        # User exists - check if active
        if not user['is_active']:
            flash('חשבון המשתמש מושבת', 'error')
            audit_logger.log_login(username, False, 'User account disabled')
            return redirect(url_for('login'))
        
        # Sync user info from Azure AD if configured
        if db.get_system_setting('ad_sync_on_login') == '1':
            ad_service.sync_user_to_local(
                ad_user_info,
                user['role'],
                user.get('hativa_id')
            )
        
        user_id = user['user_id']
        role = user['role']
        hativa_id = user.get('hativa_id')
    else:
        # New Azure AD user - auto-create if configured
        auto_create = db.get_system_setting('ad_auto_create_users')
        app.logger.info(f"User not found. Auto-create setting: {auto_create}")
        
        if auto_create == '1':
            # Determine role from Azure AD groups
            role = ad_service.get_default_role_from_groups(ad_user_info.get('groups', []))
            app.logger.info(f"Determined role: {role}")
            
            # Get default hativa
            default_hativa_str = db.get_system_setting('ad_default_hativa_id')
            default_hativa_id = int(default_hativa_str) if default_hativa_str else None
            app.logger.info(f"Default hativa_id: {default_hativa_id}")
            
            # Create user
            app.logger.info(f"Creating new user: {username}")
            try:
                user_id = ad_service.sync_user_to_local(
                    ad_user_info,
                    role,
                    default_hativa_id
                )
                app.logger.info(f"User created with ID: {user_id}")
            except Exception as e:
                app.logger.error(f"Error creating user: {e}", exc_info=True)
                flash(f'שגיאה ביצירת חשבון: {str(e)}', 'error')
                return redirect(url_for('login'))
            
            if not user_id:
                app.logger.error("User creation failed - sync_user_to_local returned None")
                flash('שגיאה ביצירת חשבון משתמש', 'error')
                return redirect(url_for('login'))
            
            hativa_id = default_hativa_id
        else:
            app.logger.warning(f"User {username} not authorized - auto-create disabled")
            flash('משתמש לא מורשה להתחבר למערכת', 'error')
            return redirect(url_for('login'))
    
    # Create session for Azure AD user
    session['user_id'] = user_id
    session['username'] = username
    session['role'] = role
    session['hativa_id'] = hativa_id
    session['full_name'] = ad_user_info['full_name']
    session['auth_source'] = 'azure_ad'
    
    # Update last login
    db.update_last_login(user_id)
    
    # Log successful login
    audit_logger.log_login(username, True, None)
    
    flash(f"ברוך הבא, {ad_user_info['full_name']}", 'success')
    return redirect(url_for('index'))


# Admin API endpoints
@app.route('/api/toggle_editing_period', methods=['POST'])
@admin_required
def toggle_editing_period():
    """Toggle editing period for regular users"""
    try:
        current_status = db.get_system_setting('editing_period_active')
        new_status = '0' if current_status == '1' else '1'
        
        user_id = session['user_id']
        db.update_system_setting('editing_period_active', new_status, user_id)
        
        status_text = "פעילה" if new_status == '1' else "סגורה"
        return jsonify({
            'success': True,
            'message': f'תקופת העריכה הכללית עכשיו {status_text}',
            'editing_active': new_status == '1'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/toggle_deadline_dates', methods=['POST'])
@login_required
def toggle_deadline_dates():
    """Toggle display of deadline dates in calendar"""
    try:
        current_status = db.get_system_setting('show_deadline_dates_in_calendar')
        new_status = '0' if current_status == '1' else '1'
        
        user_id = session['user_id']
        db.update_system_setting('show_deadline_dates_in_calendar', new_status, user_id)
        
        status_text = "מוצגים" if new_status == '1' else "מוסתרים"
        return jsonify({
            'success': True,
            'message': f'תאריכי דדליין עכשיו {status_text}',
            'show_deadline_dates': new_status == '1'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/editing_status')
@login_required
def get_editing_status():
    """Get current editing status for user"""
    try:
        user = auth_manager.get_current_user()
        can_edit, reason = auth_manager.can_edit()
        editing_period_active = db.get_system_setting('editing_period_active') == '1'
        
        return jsonify({
            'can_edit': can_edit,
            'reason': reason,
            'editing_period_active': editing_period_active,
            'user_role': user['role']
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Bulk Delete API endpoints
@app.route('/api/events/bulk_delete', methods=['POST'])
@login_required
def bulk_delete_events():
    """Bulk delete events by IDs"""
    try:
        data = request.get_json(silent=True) or {}
        event_ids = data.get('event_ids') or []
        if not isinstance(event_ids, list) or not all(isinstance(x, (int, str)) for x in event_ids):
            return jsonify({'success': False, 'message': 'פורמט בקשה לא תקין'}), 400

        # Optional: limit batch size
        if len(event_ids) == 0:
            return jsonify({'success': True, 'deleted_count': 0, 'message': 'לא נבחרו אירועים למחיקה'})
        if len(event_ids) > 1000:
            return jsonify({'success': False, 'message': 'ניתן למחוק עד 1000 אירועים בבת אחת'}), 400

        # Get current user
        user = auth_manager.get_current_user()
        if not user:
            return jsonify({'success': False, 'message': 'נדרשת התחברות'}), 401
        
        # Check permissions: Users cannot delete events
        if user['role'] == 'user':
            return jsonify({'success': False, 'message': 'משתמשים רגילים לא יכולים למחוק אירועים'}), 403
        
        # Pre-fetch names for audit per item (best-effort)
        events_map = {e['event_id']: e for e in db.get_all_events()}
        
        # For managers, verify all events are in their division
        if user['role'] == 'manager':
            for event_id in event_ids:
                try:
                    event_id_int = int(event_id)
                    if event_id_int in events_map:
                        event = events_map[event_id_int]
                        # Check if event belongs to manager's division
                        if event.get('maslul_hativa_id') != user['hativa_id']:
                            return jsonify({
                                'success': False, 
                                'message': f'מנהל יכול למחוק רק אירועים בחטיבה שלו. אירוע {event.get("name", event_id)} לא בחטיבה שלך'
                            }), 403
                except (ValueError, KeyError):
                    continue
        deleted = db.delete_events_bulk(event_ids, user['user_id'])

        # Audit: summary
        audit_logger.log_success(
            audit_logger.ACTION_DELETE,
            audit_logger.ENTITY_EVENT,
            details=f'מחיקה מרובה: {deleted} אירועים'
        )
        # Audit: per-item (only for those that existed)
        for eid in event_ids:
            try:
                eid_int = int(eid)
                if eid_int in events_map:
                    audit_logger.log_event_deleted(eid_int, events_map[eid_int].get('name', 'Unknown'))
            except Exception:
                continue

        return jsonify({'success': True, 'deleted_count': deleted})
    except Exception as e:
        audit_logger.log_error(
            audit_logger.ACTION_DELETE,
            audit_logger.ENTITY_EVENT,
            str(e),
            details='מחיקה מרובה אירועים'
        )
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/committees/bulk_delete', methods=['POST'])
@login_required
def bulk_delete_committees():
    """Bulk delete committees (vaadot) by IDs; related events are removed via cascade."""
    try:
        data = request.get_json(silent=True) or {}
        vaadot_ids = data.get('vaadot_ids') or []
        if not isinstance(vaadot_ids, list) or not all(isinstance(x, (int, str)) for x in vaadot_ids):
            return jsonify({'success': False, 'message': 'פורמט בקשה לא תקין'}), 400

        if len(vaadot_ids) == 0:
            return jsonify({'success': True, 'deleted_committees': 0, 'deleted_events': 0, 'message': 'לא נבחרו ועדות למחיקה'})
        if len(vaadot_ids) > 500:
            return jsonify({'success': False, 'message': 'ניתן למחוק עד 500 ועדות בבת אחת'}), 400

        # Get current user
        user = auth_manager.get_current_user()
        if not user:
            return jsonify({'success': False, 'message': 'נדרשת התחברות'}), 401
        
        # Check permissions: Users cannot delete committees
        if user['role'] == 'user':
            return jsonify({'success': False, 'message': 'משתמשים רגילים לא יכולים למחוק ועדות'}), 403
        
        # Pre-fetch names for audit per item (best-effort)
        vaadot_map = {v['vaadot_id']: v for v in db.get_vaadot()}
        
        # For managers, verify all committees are in their division
        if user['role'] == 'manager':
            for vaada_id in vaadot_ids:
                try:
                    vaada_id_int = int(vaada_id)
                    if vaada_id_int in vaadot_map:
                        vaada = vaadot_map[vaada_id_int]
                        # Check if committee belongs to manager's division
                        if vaada.get('hativa_id') != user['hativa_id']:
                            return jsonify({
                                'success': False, 
                                'message': f'מנהל יכול למחוק רק ועדות בחטיבה שלו. ועדה {vaada.get("committee_name", vaada_id)} לא בחטיבה שלך'
                            }), 403
                except (ValueError, KeyError):
                    continue
        deleted_committees, affected_events = db.delete_vaadot_bulk(vaadot_ids, user['user_id'])

        # Audit: summary
        audit_logger.log_success(
            audit_logger.ACTION_DELETE,
            audit_logger.ENTITY_VAADA,
            details=f'מחיקה מרובה: {deleted_committees} ועדות ו-{affected_events} אירועים'
        )
        # Audit: per-item
        for vid in vaadot_ids:
            try:
                vid_int = int(vid)
                committee_name = vaadot_map.get(vid_int, {}).get('committee_name', 'Unknown')
                audit_logger.log_vaada_deleted(vid_int, committee_name)
            except Exception:
                continue

        return jsonify({'success': True, 'deleted_committees': deleted_committees, 'deleted_events': affected_events})
    except Exception as e:
        audit_logger.log_error(
            audit_logger.ACTION_DELETE,
            audit_logger.ENTITY_VAADA,
            str(e),
            details='מחיקה מרובה ועדות'
        )
        return jsonify({'success': False, 'message': str(e)}), 500

# Soft Delete API endpoints
@app.route('/api/toggle_committee_type/<int:committee_type_id>', methods=['POST'])
@editing_permission_required
def toggle_committee_type(committee_type_id):
    """Toggle committee type active status"""
    try:
        # Check current status
        committee_types = db.get_committee_types()
        committee_type = next((ct for ct in committee_types if ct['committee_type_id'] == committee_type_id), None)
        
        if not committee_type:
            return jsonify({'success': False, 'error': 'סוג ועדה לא נמצא'}), 404
        
        # Toggle status
        if committee_type.get('is_active', 1):
            success = db.deactivate_committee_type(committee_type_id)
            action = 'הושבת'
        else:
            success = db.activate_committee_type(committee_type_id)
            action = 'הופעל'
        
        if success:
            return jsonify({
                'success': True,
                'message': f'סוג הועדה "{committee_type["name"]}" {action} בהצלחה',
                'is_active': not committee_type.get('is_active', 1)
            })
        else:
            return jsonify({'success': False, 'error': 'שגיאה בעדכון סטטוס'}), 500
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/toggle_hativa/<int:hativa_id>', methods=['POST'])
@admin_required
def toggle_hativa(hativa_id):
    """Toggle division active status (admin only)"""
    try:
        hativot = db.get_hativot()
        hativa = next((h for h in hativot if h['hativa_id'] == hativa_id), None)
        
        if not hativa:
            return jsonify({'success': False, 'error': 'חטיבה לא נמצאה'}), 404
        
        # Toggle status
        if hativa.get('is_active', 1):
            success = db.deactivate_hativa(hativa_id)
            action = 'הושבתה'
        else:
            success = db.activate_hativa(hativa_id)
            action = 'הופעלה'
        
        if success:
            return jsonify({
                'success': True,
                'message': f'החטיבה "{hativa["name"]}" {action} בהצלחה',
                'is_active': not hativa.get('is_active', 1)
            })
        else:
            return jsonify({'success': False, 'error': 'שגיאה בעדכון סטטוס'}), 500
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/toggle_maslul/<int:maslul_id>', methods=['POST'])
@editing_permission_required
def toggle_maslul(maslul_id):
    """Toggle route active status"""
    try:
        maslulim = db.get_maslulim()
        maslul = next((m for m in maslulim if m['maslul_id'] == maslul_id), None)
        
        if not maslul:
            return jsonify({'success': False, 'error': 'מסלול לא נמצא'}), 404
        
        # Toggle status
        if maslul.get('is_active', 1):
            success = db.deactivate_maslul(maslul_id)
            action = 'הושבת'
        else:
            success = db.activate_maslul(maslul_id)
            action = 'הופעל'
        
        if success:
            return jsonify({
                'success': True,
                'message': f'המסלול "{maslul["name"]}" {action} בהצלחה',
                'is_active': not maslul.get('is_active', 1)
            })
        else:
            return jsonify({'success': False, 'error': 'שגיאה בעדכון סטטוס'}), 500
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Enhanced SLA and Business Days API
@app.route('/api/sla_info/<vaadot_id>')
@login_required
def get_sla_info(vaadot_id):
    """Get SLA information for a committee meeting"""
    try:
        # Get committee meeting details
        committees = db.get_vaadot()
        committee = next((c for c in committees if c['vaadot_id'] == int(vaadot_id)), None)
        
        if not committee:
            return jsonify({'error': 'ישיבת ועדה לא נמצאה'}), 404
        
        # Parse committee date
        committee_date = datetime.strptime(committee['vaada_date'], '%Y-%m-%d').date()
        
        # Calculate SLA dates
        sla_info = db.calculate_sla_dates(committee_date)
        
        # Format dates for display
        formatted_sla = {
            'committee_date': committee_date.strftime('%d/%m/%Y'),
            'committee_name': committee['committee_name'],
            'hativa_name': committee['hativa_name'],
            'sla_days': sla_info['sla_days'],
            'request_deadline': sla_info['request_deadline'].strftime('%d/%m/%Y'),
            'preparation_start': sla_info['preparation_start'].strftime('%d/%m/%Y'),
            'notification_date': sla_info['notification_date'].strftime('%d/%m/%Y'),
            'business_days_to_committee': sla_info['business_days_to_committee'],
            'is_overdue': sla_info['request_deadline'] < date.today(),
            'days_until_deadline': (sla_info['request_deadline'] - date.today()).days
        }
        
        return jsonify(formatted_sla)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/business_days/<int:year>/<int:month>')
@login_required
def get_business_days_info(year, month):
    """Get business days information for a specific month"""
    try:
        business_days_info = db.get_monthly_business_days(year, month)
        
        # Format dates for JSON
        formatted_info = {
            'year': business_days_info['year'],
            'month': business_days_info['month'],
            'total_days': business_days_info['total_days'],
            'business_days_count': business_days_info['business_days_count'],
            'business_days': [d.strftime('%Y-%m-%d') for d in business_days_info['business_days']],
            'first_business_day': business_days_info['first_business_day'].strftime('%d/%m/%Y') if business_days_info['first_business_day'] else None,
            'last_business_day': business_days_info['last_business_day'].strftime('%d/%m/%Y') if business_days_info['last_business_day'] else None,
            'weeks': {
                str(week): [d.strftime('%Y-%m-%d') for d in days] 
                for week, days in business_days_info['weeks'].items()
            }
        }
        
        return jsonify(formatted_info)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/available_dates')
@login_required
def get_available_dates():
    """Get upcoming available dates for a committee type within a division"""
    try:
        committee_type_id = request.args.get('committee_type_id', type=int)
        hativa_id = request.args.get('hativa_id', type=int)

        if not committee_type_id or not hativa_id:
            return jsonify({'success': False, 'message': 'נדרש לבחור סוג ועדה וחטיבה'}), 400

        # Validate that the committee type belongs to the given division
        committee_types = db.get_committee_types(hativa_id)
        committee_type = next((ct for ct in committee_types if ct['committee_type_id'] == committee_type_id), None)

        if not committee_type:
            return jsonify({'success': False, 'message': 'סוג הועדה אינו משויך לחטיבה שנבחרה'}), 400

        start_date_str = request.args.get('start_date')
        start_date = None
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'success': False, 'message': 'פורמט תאריך התחלה אינו תקין'}), 400

        limit = request.args.get('limit', default=5, type=int)
        limit = max(1, min(limit or 5, 20))
        max_days = request.args.get('max_days', default=180, type=int)
        max_days = max(7, min(max_days or 180, 365))

        available_dates = auto_scheduler.find_available_dates(
            committee_type_id,
            hativa_id,
            start_date=start_date,
            max_results=limit,
            max_days=max_days
        )

        formatted_dates = [d.strftime('%Y-%m-%d') for d in available_dates]

        message = 'נמצאו תאריכים פנויים' if formatted_dates else 'לא נמצאו תאריכים פנויים בטווח שנבחר'

        return jsonify({
            'success': True,
            'dates': formatted_dates,
            'total': len(formatted_dates),
            'message': message,
            'committee_type': committee_type.get('name'),
            'hativa_id': hativa_id
        })

    except Exception as e:
        app.logger.error(f"Error fetching available dates: {str(e)}")
        return jsonify({'success': False, 'message': f'שגיאה בשליפת תאריכים פנויים: {str(e)}'}), 500


@app.route('/api/recommend_committees')
@login_required
def recommend_committees():
    """Get committee recommendations for a new event based on route and expected requests"""
    try:
        maslul_id = request.args.get('maslul_id', type=int)
        expected_requests = request.args.get('expected_requests', default=0, type=int)
        event_name = request.args.get('event_name', default='', type=str)
        limit = request.args.get('limit', default=5, type=int)
        
        if not maslul_id:
            return jsonify({'success': False, 'message': 'נדרש לבחור מסלול'}), 400
        
        # Get recommendations
        recommendations = committee_recommendation_service.recommend_committees(
            maslul_id=maslul_id,
            expected_requests=expected_requests,
            event_name=event_name,
            limit=min(limit, 10)  # Max 10 recommendations
        )
        
        return jsonify({
            'success': True,
            'recommendations': recommendations,
            'total': len(recommendations)
        })
        
    except Exception as e:
        import traceback
        app.logger.error(f"Error getting committee recommendations: {str(e)}")
        app.logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': f'שגיאה בקבלת המלצות: {str(e)}'}), 500


@app.route('/constraints')
@admin_required
def constraints():
    """Constraint management dashboard"""
    try:
        settings = constraints_service.get_constraints_overview()
        current_user = auth_manager.get_current_user()
        return render_template('constraints.html', settings=settings, errors={}, current_user=current_user)
    except Exception as e:
        flash(f'שגיאה בטעינת נתוני האילוצים: {e}', 'error')
        return redirect(url_for('index'))


@app.route('/constraints/update', methods=['POST'])
@admin_required
def update_constraints():
    """Update system constraint settings"""
    try:
        payload = constraints_service.parse_request(request.form)
        result = constraints_service.update_constraints(payload, session.get('user_id'))
        if result.success:
            flash(result.message, 'success')
            return redirect(url_for('constraints'))
        flash(result.message, 'error')
        settings = constraints_service.get_constraints_overview()
        settings = constraints_service.apply_form_values(settings, payload)
        current_user = auth_manager.get_current_user()
        status_code = 400 if result.errors else 200
        return render_template('constraints.html', settings=settings, errors=result.errors or {}, current_user=current_user), status_code
    except Exception as e:
        flash(f'שגיאה בעדכון האילוצים: {e}', 'error')
        return redirect(url_for('constraints'))

@app.route('/')
@login_required
def index():
    """Main dashboard"""
    # Get summary statistics
    hativot = db.get_hativot()
    maslulim = db.get_maslulim()
    committee_types = db.get_committee_types()
    committees = db.get_vaadot()  # This now returns meeting instances
    events = db.get_all_events()
    exception_dates = db.get_exception_dates()
    
    # Debug logging
    app.logger.info(f"Loaded {len(committees)} committees")
    app.logger.info(f"Loaded {len(events)} events")
    if committees:
        app.logger.info(f"First committee: {committees[0]}")
    if events:
        app.logger.info(f"First event: {events[0]}")
    
    # Get current month schedule
    today = date.today()
    monthly_schedule = []
    
    stats = {
        'hativot_count': len(hativot),
        'maslulim_count': len(maslulim),
        'committee_types_count': len(committee_types),
        'committees_count': len(committees),
        'events_count': len(events),
        'exception_dates_count': len(exception_dates),
        'business_days_this_month': 0
    }
    
    # Get current user info
    current_user = auth_manager.get_current_user()
    
    # Get deadline dates display setting
    show_deadline_dates = db.get_system_setting('show_deadline_dates_in_calendar') == '1'
    
    return render_template('index.html', 
                         hativot=hativot, 
                         maslulim=maslulim, 
                         committee_types=committee_types,
                         committees=committees,
                         events=events,
                         exception_dates=exception_dates,
                         stats=stats,
                         current_user=current_user,
                         show_deadline_dates=show_deadline_dates)

@app.route('/dashboard')
@login_required
def dashboard():
    """Analytics Dashboard"""
    from datetime import datetime, timedelta
    from collections import defaultdict
    
    # Get all data
    hativot = db.get_hativot()
    maslulim = db.get_maslulim()
    committee_types = db.get_committee_types()
    committees = db.get_vaadot()
    events = db.get_all_events()
    
    # Current date
    today = date.today()
    current_month = today.month
    current_year = today.year
    
    # === Statistics by Division ===
    stats_by_hativa = {}
    for hativa in hativot:
        hativa_id = hativa['hativa_id']
        hativa_events = [e for e in events if e.get('hativa_id') == hativa_id]
        hativa_committees = [c for c in committees if c.get('hativa_id') == hativa_id]
        
        total_expected = sum([e.get('expected_requests', 0) or 0 for e in hativa_events])
        total_actual = sum([e.get('actual_submissions', 0) or 0 for e in hativa_events])
        
        stats_by_hativa[hativa_id] = {
            'name': hativa['name'],
            'color': hativa.get('color', '#007bff'),
            'events_count': len(hativa_events),
            'committees_count': len(hativa_committees),
            'expected_requests': total_expected,
            'actual_submissions': total_actual,
            'fulfillment_rate': round((total_actual / total_expected * 100) if total_expected > 0 else 0, 1)
        }
    
    # === Events by Type ===
    events_by_type = defaultdict(int)
    for event in events:
        event_type = event.get('event_type', 'אחר')
        events_by_type[event_type] += 1
    
    # === Monthly Trend (last 6 months) ===
    monthly_data = {}
    for i in range(6):
        month_date = today - timedelta(days=30 * i)
        month_key = f"{month_date.year}-{month_date.month:02d}"
        monthly_data[month_key] = {
            'committees': 0,
            'events': 0,
            'expected_requests': 0,
            'actual_submissions': 0
        }
    
    for committee in committees:
        if committee.get('vaada_date'):
            try:
                if isinstance(committee['vaada_date'], str):
                    vaada_date = datetime.strptime(committee['vaada_date'], '%Y-%m-%d').date()
                else:
                    vaada_date = committee['vaada_date']
                
                month_key = f"{vaada_date.year}-{vaada_date.month:02d}"
                if month_key in monthly_data:
                    monthly_data[month_key]['committees'] += 1
            except:
                pass
    
    for event in events:
        if event.get('committee_date'):
            try:
                if isinstance(event['committee_date'], str):
                    event_date = datetime.strptime(event['committee_date'], '%Y-%m-%d').date()
                else:
                    event_date = event['committee_date']
                
                month_key = f"{event_date.year}-{event_date.month:02d}"
                if month_key in monthly_data:
                    monthly_data[month_key]['events'] += 1
                    monthly_data[month_key]['expected_requests'] += event.get('expected_requests', 0) or 0
                    monthly_data[month_key]['actual_submissions'] += event.get('actual_submissions', 0) or 0
            except:
                pass
    
    # === Top Routes by Events ===
    maslul_stats = defaultdict(lambda: {'count': 0, 'expected': 0, 'actual': 0})
    for event in events:
        maslul_id = event.get('maslul_id')
        if maslul_id:
            maslul_stats[maslul_id]['count'] += 1
            maslul_stats[maslul_id]['expected'] += event.get('expected_requests', 0) or 0
            maslul_stats[maslul_id]['actual'] += event.get('actual_submissions', 0) or 0
    
    # Add maslul names
    maslul_rankings = []
    for maslul in maslulim:
        maslul_id = maslul['maslul_id']
        if maslul_id in maslul_stats:
            stats = maslul_stats[maslul_id]
            maslul_rankings.append({
                'name': maslul['name'],
                'hativa_name': maslul.get('hativa_name', ''),
                'events_count': stats['count'],
                'expected_requests': stats['expected'],
                'actual_submissions': stats['actual'],
                'fulfillment_rate': round((stats['actual'] / stats['expected'] * 100) if stats['expected'] > 0 else 0, 1)
            })
    
    maslul_rankings.sort(key=lambda x: x['events_count'], reverse=True)
    top_maslulim = maslul_rankings[:10]
    
    # === Upcoming Events (next 30 days) ===
    upcoming_events = []
    future_date = today + timedelta(days=30)
    for event in events:
        if event.get('vaada_date'):
            try:
                if isinstance(event['vaada_date'], str):
                    event_date = datetime.strptime(event['vaada_date'], '%Y-%m-%d').date()
                else:
                    event_date = event['vaada_date']
                
                if today <= event_date <= future_date:
                    upcoming_events.append(event)
            except:
                pass
    
    upcoming_events.sort(key=lambda x: x.get('vaada_date', ''))
    
    # === Overall Statistics ===
    total_expected = sum([e.get('expected_requests', 0) or 0 for e in events])
    total_actual = sum([e.get('actual_submissions', 0) or 0 for e in events])
    overall_fulfillment = round((total_actual / total_expected * 100) if total_expected > 0 else 0, 1)
    
    stats = {
        'total_hativot': len(hativot),
        'total_maslulim': len(maslulim),
        'total_committee_types': len(committee_types),
        'total_committees': len(committees),
        'total_events': len(events),
        'total_expected_requests': total_expected,
        'total_actual_submissions': total_actual,
        'overall_fulfillment_rate': overall_fulfillment,
        'kokok_count': events_by_type.get('kokok', 0),
        'shotef_count': events_by_type.get('shotef', 0)
    }
    
    current_user = auth_manager.get_current_user()
    
    return render_template('dashboard.html',
                         stats=stats,
                         stats_by_hativa=stats_by_hativa,
                         events_by_type=dict(events_by_type),
                         monthly_data=monthly_data,
                         top_maslulim=top_maslulim,
                         upcoming_events=upcoming_events[:10],
                         current_user=current_user)

@app.route('/hativot')
@login_required
def hativot():
    """Manage divisions"""
    hativot_list = db.get_hativot()
    current_user = auth_manager.get_current_user()
    return render_template('hativot.html', hativot=hativot_list, current_user=current_user)

@app.route('/hativot/add', methods=['POST'])
@editing_permission_required
def add_hativa():
    """Add new division"""
    name = request.form.get('name', '').strip()
    description = request.form.get('description', '').strip()
    color = request.form.get('color', '#007bff')
    
    if not name:
        flash('שם החטיבה הוא שדה חובה', 'error')
        return redirect(url_for('hativot'))
    
    try:
        hativa_id = db.add_hativa(name, description, color)
        audit_logger.log_hativa_created(hativa_id, name)
        flash(f'חטיבה "{name}" נוספה בהצלחה', 'success')
    except Exception as e:
        audit_logger.log_error(audit_logger.ACTION_CREATE, audit_logger.ENTITY_HATIVA, str(e), entity_name=name)
        flash(f'שגיאה בהוספת החטיבה: {str(e)}', 'error')
    
    return redirect(url_for('hativot'))

@app.route('/hativot/update', methods=['POST'])
@editing_permission_required
def update_hativa():
    """Update existing division"""
    hativa_id = request.form.get('hativa_id')
    name = request.form.get('name', '').strip()
    description = request.form.get('description', '').strip()
    color = request.form.get('color', '#007bff')
    
    if not all([hativa_id, name]):
        flash('מזהה החטיבה ושם החטיבה הם שדות חובה', 'error')
        return redirect(url_for('hativot'))
    
    try:
        success = db.update_hativa(int(hativa_id), name, description, color)
        if success:
            audit_logger.log_hativa_updated(int(hativa_id), name)
            flash(f'חטיבה "{name}" עודכנה בהצלחה', 'success')
        else:
            flash('שגיאה בעדכון החטיבה', 'error')
    except Exception as e:
        audit_logger.log_error(audit_logger.ACTION_UPDATE, audit_logger.ENTITY_HATIVA, str(e), int(hativa_id), name)
        flash(f'שגיאה בעדכון החטיבה: {str(e)}', 'error')
    
    return redirect(url_for('hativot'))

@app.route('/maslulim')
def maslulim():
    """Manage routes with enhanced functionality"""
    try:
        # Get data with error handling
        maslulim_list = db.get_maslulim()
        hativot_list = db.get_hativot()
        
        # Group maslulim by hativa for better organization and display control
        maslulim_by_hativa = []
        for hativa in hativot_list:
            grouped_maslulim = [m for m in maslulim_list if m['hativa_id'] == hativa['hativa_id']]
            maslulim_by_hativa.append({
                'hativa_id': hativa['hativa_id'],
                'hativa_name': hativa['name'],
                'color': hativa.get('color'),
                'maslulim': grouped_maslulim
            })
        
        # Calculate statistics with colors
        maslulim_per_hativa_with_colors = {}
        for hativa in hativot_list:
            count = len([m for m in maslulim_list if m['hativa_id'] == hativa['hativa_id']])
            maslulim_per_hativa_with_colors[hativa['name']] = {
                'count': count,
                'color': hativa['color']
            }
        
        stats = {
            'total_maslulim': len(maslulim_list),
            'total_hativot': len(hativot_list),
            'maslulim_per_hativa': {hativa['name']: len([m for m in maslulim_list if m['hativa_id'] == hativa['hativa_id']]) for hativa in hativot_list},
            'maslulim_per_hativa_with_colors': maslulim_per_hativa_with_colors
        }
        
        # Get current user info
        current_user = auth_manager.get_current_user()
        
        return render_template('maslulim.html', 
                             maslulim=maslulim_list, 
                             hativot=hativot_list,
                             maslulim_by_hativa=maslulim_by_hativa,
                             stats=stats,
                             current_user=current_user)
    except Exception as e:
        flash(f'שגיאה בטעינת נתוני המסלולים: {str(e)}', 'error')
        current_user = auth_manager.get_current_user()
        return render_template('maslulim.html', maslulim=[], hativot=[], maslulim_by_hativa={}, stats={}, current_user=current_user)

@app.route('/maslulim/add', methods=['POST'])
def add_maslul():
    """Add new route with enhanced validation"""
    try:
        hativa_id = request.form.get('hativa_id')
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        sla_days = request.form.get('sla_days', 45)
        stage_a_days = request.form.get('stage_a_days', 10)
        stage_b_days = request.form.get('stage_b_days', 15)
        stage_c_days = request.form.get('stage_c_days', 10)
        stage_d_days = request.form.get('stage_d_days', 10)
        
        # Enhanced validation
        if not hativa_id:
            flash('יש לבחור חטיבה', 'error')
            return redirect(url_for('maslulim'))
            
        if not name:
            flash('שם המסלול הוא שדה חובה', 'error')
            return redirect(url_for('maslulim'))
            
        if len(name) < 2:
            flash('שם המסלול חייב להכיל לפחות 2 תווים', 'error')
            return redirect(url_for('maslulim'))
            
        # Check if hativa exists
        hativot = db.get_hativot()
        if not any(h['hativa_id'] == int(hativa_id) for h in hativot):
            flash('החטיבה שנבחרה לא קיימת במערכת', 'error')
            return redirect(url_for('maslulim'))
            
        # Check for duplicate names within the same hativa
        existing_maslulim = db.get_maslulim(int(hativa_id))
        if any(m['name'].lower() == name.lower() for m in existing_maslulim):
            flash(f'מסלול בשם "{name}" כבר קיים בחטיבה זו', 'error')
            return redirect(url_for('maslulim'))
        
        # Validate SLA and stage days
        try:
            sla_days = int(sla_days)
            stage_a_days = int(stage_a_days)
            stage_b_days = int(stage_b_days)
            stage_c_days = int(stage_c_days)
            stage_d_days = int(stage_d_days)
            
            if sla_days < 1 or sla_days > 365:
                flash('SLA חייב להיות בין 1 ל-365 ימים', 'error')
                return redirect(url_for('maslulim'))
                
            # Validate stage days
            for stage_name, days in [('שלב א', stage_a_days), ('שלב ב', stage_b_days), 
                                   ('שלב ג', stage_c_days), ('שלב ד', stage_d_days)]:
                if days < 0 or days > 365:
                    flash(f'{stage_name} חייב להיות בין 0 ל-365 ימים', 'error')
                    return redirect(url_for('maslulim'))
            
            # Check if sum of stages equals SLA
            total_stages = stage_a_days + stage_b_days + stage_c_days + stage_d_days
            if total_stages != sla_days:
                flash(f'סכום השלבים ({total_stages}) חייב להיות שווה ל-SLA ({sla_days})', 'error')
                return redirect(url_for('maslulim'))
                
        except (ValueError, TypeError):
            flash('כל השדות חייבים להיות מספרים תקינים', 'error')
            return redirect(url_for('maslulim'))
        
        # Add the maslul
        maslul_id = db.add_maslul(int(hativa_id), name, description, sla_days, 
                                 stage_a_days, stage_b_days, stage_c_days, stage_d_days)
        hativa_name = next(h['name'] for h in hativot if h['hativa_id'] == int(hativa_id))
        audit_logger.log_maslul_created(maslul_id, name, hativa_name)
        flash(f'מסלול "{name}" נוסף בהצלחה לחטיבת {hativa_name}', 'success')
        
    except ValueError as e:
        audit_logger.log_error(audit_logger.ACTION_CREATE, audit_logger.ENTITY_MASLUL, str(e), entity_name=name)
        flash('נתונים לא תקינים', 'error')
    except Exception as e:
        audit_logger.log_error(audit_logger.ACTION_CREATE, audit_logger.ENTITY_MASLUL, str(e), entity_name=name)
        flash(f'שגיאה בהוספת המסלול: {str(e)}', 'error')
    
    return redirect(url_for('maslulim'))

@app.route('/maslulim/edit/<int:maslul_id>', methods=['POST'])
def edit_maslul(maslul_id):
    """Edit existing route"""
    try:
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        sla_days = request.form.get('sla_days', 45)
        stage_a_days = request.form.get('stage_a_days', 10)
        stage_b_days = request.form.get('stage_b_days', 15)
        stage_c_days = request.form.get('stage_c_days', 10)
        stage_d_days = request.form.get('stage_d_days', 10)
        is_active = request.form.get('is_active') == 'on'
        
        if not name:
            flash('שם המסלול הוא שדה חובה', 'error')
            return redirect(url_for('maslulim'))
            
        if len(name) < 2:
            flash('שם המסלול חייב להכיל לפחות 2 תווים', 'error')
            return redirect(url_for('maslulim'))
        
        # Validate SLA and stage days
        try:
            sla_days = int(sla_days)
            stage_a_days = int(stage_a_days)
            stage_b_days = int(stage_b_days)
            stage_c_days = int(stage_c_days)
            stage_d_days = int(stage_d_days)
            
            if sla_days < 1 or sla_days > 365:
                flash('SLA חייב להיות בין 1 ל-365 ימים', 'error')
                return redirect(url_for('maslulim'))
                
            # Validate stage days
            for stage_name, days in [('שלב א', stage_a_days), ('שלב ב', stage_b_days), 
                                   ('שלב ג', stage_c_days), ('שלב ד', stage_d_days)]:
                if days < 0 or days > 365:
                    flash(f'{stage_name} חייב להיות בין 0 ל-365 ימים', 'error')
                    return redirect(url_for('maslulim'))
            
            # Check if sum of stages equals SLA
            total_stages = stage_a_days + stage_b_days + stage_c_days + stage_d_days
            if total_stages != sla_days:
                flash(f'סכום השלבים ({total_stages}) חייב להיות שווה ל-SLA ({sla_days})', 'error')
                return redirect(url_for('maslulim'))
                
        except (ValueError, TypeError):
            flash('כל השדות חייבים להיות מספרים תקינים', 'error')
            return redirect(url_for('maslulim'))
        
        # Update the maslul
        success = db.update_maslul(maslul_id, name, description, sla_days, 
                                 stage_a_days, stage_b_days, stage_c_days, stage_d_days, is_active)
        if success:
            status_text = 'פעיל' if is_active else 'לא פעיל'
            flash(f'מסלול "{name}" עודכן בהצלחה (סטטוס: {status_text})', 'success')
        else:
            flash('המסלול לא נמצא במערכת', 'error')
            
    except Exception as e:
        flash(f'שגיאה בעדכון המסלול: {str(e)}', 'error')
    
    return redirect(url_for('maslulim'))

@app.route('/maslulim/delete/<int:maslul_id>', methods=['POST'])
def delete_maslul(maslul_id):
    """Delete route with safety checks"""
    try:
        # Check if maslul is used in any events
        events = db.get_all_events()
        maslul_events = [e for e in events if e['maslul_id'] == maslul_id]
        
        if maslul_events:
            flash(f'לא ניתן למחוק מסלול המשויך ל-{len(maslul_events)} אירועים. יש למחוק תחילה את האירועים הקשורים.', 'error')
            return redirect(url_for('maslulim'))
        
        # Get maslul name before deletion
        maslulim = db.get_maslulim()
        maslul = next((m for m in maslulim if m['maslul_id'] == maslul_id), None)
        
        if not maslul:
            flash('המסלול לא נמצא במערכת', 'error')
            return redirect(url_for('maslulim'))
        
        # Delete the maslul
        success = db.delete_maslul(maslul_id)
        if success:
            flash(f'מסלול "{maslul["name"]}" נמחק בהצלחה', 'success')
        else:
            flash('שגיאה במחיקת המסלול', 'error')
            
    except Exception as e:
        flash(f'שגיאה במחיקת המסלול: {str(e)}', 'error')
    
    return redirect(url_for('maslulim'))

@app.route('/exception_dates', methods=['GET', 'POST'])
def exception_dates():
    """Manage exception dates"""
    if request.method == 'POST':
        # Handle POST - add new exception date
        date_str = request.form.get('date', '').strip()
        description = request.form.get('description', '').strip()
        date_type = request.form.get('type', 'holiday').strip()
        
        if not date_str:
            flash('תאריך הוא שדה חובה', 'error')
            return redirect(url_for('exception_dates'))
        
        try:
            exception_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            db.add_exception_date(exception_date, description, date_type)
            
            # Recalculate all event deadlines to account for the new exception date
            updated_count = db.recalculate_all_event_deadlines()
            
            # Get the date_id for logging (query just added record)
            exception_dates_list = db.get_exception_dates(include_past=True)
            added_date = next((ed for ed in exception_dates_list if ed['exception_date'] == date_str), None)
            date_id = added_date['date_id'] if added_date else None
            
            audit_logger.log_exception_date_added(date_id, date_str, description)
            flash(f'תאריך חריג {date_str} נוסף בהצלחה. עודכנו {updated_count} אירועים.', 'success')
        except ValueError:
            flash('פורמט תאריך לא תקין', 'error')
        except Exception as e:
            flash(f'שגיאה בהוספת תאריך: {str(e)}', 'error')
        
        return redirect(url_for('exception_dates'))
    
    # Handle GET - display exception dates
    include_past = request.args.get('include_past', 'true') == 'true'
    dates_list = db.get_exception_dates(include_past=include_past)
    current_user = auth_manager.get_current_user()
    return render_template('exception_dates.html', dates=dates_list, current_user=current_user, include_past=include_past)

@app.route('/exception_dates/edit/<int:date_id>', methods=['POST'])
@login_required
def edit_exception_date(date_id):
    """Edit exception date"""
    try:
        # Get current user
        user = auth_manager.get_current_user()
        if not user:
            flash('נדרשת התחברות', 'error')
            return redirect(url_for('login'))
        
        # Check permissions: Only admins can edit
        if user['role'] != 'admin':
            flash('רק מנהלי מערכת יכולים לערוך תאריכי חריגים', 'error')
            return redirect(url_for('exception_dates'))
        
        date_str = request.form.get('date', '').strip()
        description = request.form.get('description', '').strip()
        date_type = request.form.get('type', 'holiday').strip()
        
        if not date_str:
            flash('תאריך הוא שדה חובה', 'error')
            return redirect(url_for('exception_dates'))
        
        exception_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        success = db.update_exception_date(date_id, exception_date, description, date_type)
        
        if success:
            # Recalculate all event deadlines to account for the updated exception date
            updated_count = db.recalculate_all_event_deadlines()
            
            audit_logger.log_success(
                audit_logger.ACTION_UPDATE,
                'exception_date',
                details=f'עדכון תאריך חריג: {date_str} - {description}'
            )
            flash(f'תאריך חריג {date_str} עודכן בהצלחה. עודכנו {updated_count} אירועים.', 'success')
        else:
            flash('שגיאה בעדכון תאריך חריג', 'error')
    except ValueError:
        flash('פורמט תאריך לא תקין', 'error')
    except Exception as e:
        flash(f'שגיאה בעדכון תאריך: {str(e)}', 'error')
    
    return redirect(url_for('exception_dates'))

@app.route('/exception_dates/delete/<int:date_id>', methods=['POST'])
@login_required
def delete_exception_date(date_id):
    """Delete exception date"""
    try:
        # Get current user
        user = auth_manager.get_current_user()
        if not user:
            flash('נדרשת התחברות', 'error')
            return redirect(url_for('login'))
        
        # Check permissions: Only admins can delete
        if user['role'] != 'admin':
            flash('רק מנהלי מערכת יכולים למחוק תאריכי חריגים', 'error')
            return redirect(url_for('exception_dates'))
        
        # Get date info before deletion for logging
        exception_date_obj = db.get_exception_date_by_id(date_id)
        
        if not exception_date_obj:
            flash('תאריך חריג לא נמצא במערכת', 'error')
            return redirect(url_for('exception_dates'))
        
        success = db.delete_exception_date(date_id)
        
        if success:
            # Recalculate all event deadlines to account for the deleted exception date
            updated_count = db.recalculate_all_event_deadlines()
            
            audit_logger.log_success(
                audit_logger.ACTION_DELETE,
                'exception_date',
                details=f'מחיקת תאריך חריג: {exception_date_obj["exception_date"]} - {exception_date_obj.get("description", "")}'
            )
            flash(f'תאריך חריג {exception_date_obj["exception_date"]} נמחק בהצלחה. עודכנו {updated_count} אירועים.', 'success')
        else:
            flash('לא ניתן למחוק תאריך חריג שמשויכות אליו ועדות', 'error')
    except Exception as e:
        flash(f'שגיאה במחיקת תאריך: {str(e)}', 'error')
    
    return redirect(url_for('exception_dates'))


@app.route('/committees/add', methods=['POST'])
def add_committee_meeting():
    """Add new committee meeting"""
    committee_type_id = request.form.get('committee_type_id')
    hativa_id = request.form.get('hativa_id')
    vaada_date = request.form.get('vaada_date')
    status = request.form.get('status', 'planned')
    notes = request.form.get('notes', '').strip()
    
    if not all([committee_type_id, hativa_id, vaada_date]):
        flash('סוג ועדה, חטיבה ותאריך הם שדות חובה', 'error')
        return redirect(url_for('index'))
    
    # Debug logging
    print(f"DEBUG: Received date value: '{vaada_date}' (type: {type(vaada_date).__name__})")
    
    try:
        # Try multiple date formats to handle different browser formats
        meeting_date = None
        date_formats = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d']
        
        for fmt in date_formats:
            try:
                meeting_date = datetime.strptime(vaada_date, fmt).date()
                break
            except ValueError:
                continue
        
        if meeting_date is None:
            raise ValueError(f'פורמט תאריך לא תקין: {vaada_date}. נא להזין תאריך בפורמט YYYY-MM-DD')
        
        # Check if user is admin for constraint override
        current_user = auth_manager.get_current_user()
        is_admin = current_user and current_user.get('role') == 'admin'
        
        # Try to add meeting (admins get warnings instead of errors)
        vaadot_id, warning_message = db.add_vaada(
            int(committee_type_id), 
            int(hativa_id), 
            meeting_date, 
            notes=notes, 
            override_constraints=is_admin
        )
        
        # Get committee name for logging
        committee_types = db.get_committee_types()
        committee_type = next((ct for ct in committee_types if ct['committee_type_id'] == int(committee_type_id)), None)
        committee_name = committee_type['name'] if committee_type else 'Unknown'
        
        audit_logger.log_vaada_created(vaadot_id, committee_name, meeting_date.strftime('%Y-%m-%d'))
        
        # Show success message with any warnings
        if warning_message:
            flash(f'ישיבת ועדה נוספה בהצלחה. {warning_message}', 'warning')
        else:
            flash('ישיבת ועדה נוספה בהצלחה', 'success')
    except ValueError as e:
        # Check if it's a date format error or our constraint error
        if "כבר קיימת ועדה בתאריך" in str(e):
            audit_logger.log_error(audit_logger.ACTION_CREATE, audit_logger.ENTITY_VAADA, str(e), details=f'תאריך: {vaada_date}')
            flash(str(e), 'error')
        else:
            audit_logger.log_error(audit_logger.ACTION_CREATE, audit_logger.ENTITY_VAADA, str(e), details=f'תאריך: {vaada_date}')
            flash(str(e), 'error')
    except Exception as e:
        audit_logger.log_error(audit_logger.ACTION_CREATE, audit_logger.ENTITY_VAADA, str(e), details=f'תאריך: {vaada_date}')
        flash(f'שגיאה בהוספת הישיבה: {str(e)}', 'error')
    
    return redirect(url_for('index'))

@app.route('/committees/edit/<int:vaadot_id>', methods=['POST'])
def edit_committee_meeting(vaadot_id):
    """Edit existing committee meeting"""
    committee_type_id = request.form.get('committee_type_id')
    hativa_id = request.form.get('hativa_id')
    vaada_date = request.form.get('vaada_date')
    notes = request.form.get('notes', '').strip()
    
    if not all([committee_type_id, hativa_id, vaada_date]):
        flash('סוג ועדה, חטיבה ותאריך הם שדות חובה', 'error')
        return redirect(url_for('index'))
    
    try:
        # Try multiple date formats to handle different browser formats
        meeting_date = None
        date_formats = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d']
        
        for fmt in date_formats:
            try:
                meeting_date = datetime.strptime(vaada_date, fmt).date()
                break
            except ValueError:
                continue
        
        if meeting_date is None:
            raise ValueError(f'פורמט תאריך לא תקין: {vaada_date}. נא להזין תאריך בפורמט YYYY-MM-DD')
        
        success = db.update_vaada(vaadot_id, int(committee_type_id), int(hativa_id), meeting_date, notes=notes)
        if success:
            # Get committee name for logging
            committee_types = db.get_committee_types()
            committee_type = next((ct for ct in committee_types if ct['committee_type_id'] == int(committee_type_id)), None)
            committee_name = committee_type['name'] if committee_type else 'Unknown'
            
            audit_logger.log_vaada_updated(vaadot_id, committee_name, meeting_date.strftime('%Y-%m-%d'))
            flash('ישיבת הועדה עודכנה בהצלחה', 'success')
        else:
            flash('שגיאה בעדכון הישיבה', 'error')
    except ValueError as e:
        audit_logger.log_error(audit_logger.ACTION_UPDATE, audit_logger.ENTITY_VAADA, str(e), vaadot_id)
        flash(str(e), 'error')
    except Exception as e:
        audit_logger.log_error(audit_logger.ACTION_UPDATE, audit_logger.ENTITY_VAADA, str(e), vaadot_id)
        flash(f'שגיאה בעדכון הישיבה: {str(e)}', 'error')
    
    return redirect(url_for('index'))

@app.route('/committees/delete/<int:vaadot_id>', methods=['POST'])
def delete_committee_meeting(vaadot_id):
    """Delete committee meeting and its events"""
    try:
        # Get committee details before deletion for logging
        vaada = db.get_vaada_by_id(vaadot_id)
        committee_name = vaada['committee_name'] if vaada else 'Unknown'
        
        # Get current user
        user = auth_manager.get_current_user()
        user_id = user['user_id'] if user else None
        
        # First delete all related events
        events = db.get_all_events()
        related_events = [e for e in events if e['vaadot_id'] == vaadot_id]
        
        for event in related_events:
            db.delete_event(event['event_id'], user_id)
            audit_logger.log_event_deleted(event['event_id'], event['name'])
        
        # Then delete the committee meeting
        success = db.delete_vaada(vaadot_id, user_id)
        if success:
            audit_logger.log_vaada_deleted(vaadot_id, committee_name)
            flash(f'ישיבת הועדה ו-{len(related_events)} אירועים הועברו לסל המחזור בהצלחה', 'success')
        else:
            flash('שגיאה במחיקת הישיבה', 'error')
    except Exception as e:
        audit_logger.log_error(audit_logger.ACTION_DELETE, audit_logger.ENTITY_VAADA, str(e), vaadot_id)
        flash(f'שגיאה במחיקת הישיבה: {str(e)}', 'error')
    
    return redirect(url_for('index'))


@app.route('/events/add', methods=['POST'])
def add_event():
    """Add new event"""
    vaadot_id = request.form.get('vaadot_id')
    maslul_id = request.form.get('maslul_id')
    name = request.form.get('name', '').strip()
    event_type = request.form.get('event_type')
    expected_requests = request.form.get('expected_requests', '0')
    actual_submissions = request.form.get('actual_submissions', '0')
    call_publication_date = request.form.get('call_publication_date')
    
    # Get manual call deadline parameters
    is_call_deadline_manual = request.form.get('is_call_deadline_manual') == '1'
    manual_call_deadline_date = request.form.get('manual_call_deadline_date')
    
    if not all([vaadot_id, maslul_id, name, event_type]):
        flash('כל השדות הם שדות חובה', 'error')
        return redirect(url_for('index'))
    
    try:
        expected_requests = int(expected_requests) if expected_requests else 0
        actual_submissions = int(actual_submissions) if actual_submissions else 0
        
        # Validate expected_requests is required and greater than 0
        if expected_requests <= 0:
            flash('מספר הבקשות הצפויות חייב להיות גדול מ-0', 'error')
            return redirect(url_for('index'))
        
        # Validate that committee and route are from the same division
        vaada = db.get_vaadot()[0] if db.get_vaadot() else None
        for v in db.get_vaadot():
            if v['vaadot_id'] == int(vaadot_id):
                vaada = v
                break
        
        maslul = None
        for m in db.get_maslulim():
            if m['maslul_id'] == int(maslul_id):
                maslul = m
                break
        
        if not vaada or not maslul:
            flash('ועדה או מסלול לא נמצאו במערכת', 'error')
            return redirect(url_for('index'))
        
        if vaada['hativa_id'] != maslul['hativa_id']:
            flash(f'שגיאה: המסלול "{maslul["name"]}" מחטיבת "{maslul["hativa_name"]}" אינו יכול להיות משויך לועדה מחטיבת "{vaada["hativa_name"]}"', 'error')
            return redirect(url_for('index'))
        
        # Check if maslul is active
        if not maslul.get('is_active', True):
            flash(f'שגיאה: המסלול "{maslul["name"]}" אינו פעיל. לא ניתן ליצור אירועים חדשים למסלול זה.', 'error')
            return redirect(url_for('index'))
        
        # Validate event data
        event_data = {
            'vaadot_id': int(vaadot_id),
            'maslul_id': int(maslul_id),
            'name': name,
            'event_type': event_type,
            'expected_requests': expected_requests,
            'call_publication_date': call_publication_date
        }
        
        event_id = db.add_event(int(vaadot_id), int(maslul_id), name, event_type, expected_requests, actual_submissions, call_publication_date,
                                 is_call_deadline_manual, manual_call_deadline_date)
        
        # Get committee name for logging
        vaada = db.get_vaada_by_id(int(vaadot_id))
        committee_name = vaada['committee_name'] if vaada else 'Unknown'
        
        audit_logger.log_event_created(event_id, name, committee_name)
        flash(f'אירוע "{name}" נוצר בהצלחה', 'success')
        
    except Exception as e:
        audit_logger.log_error(audit_logger.ACTION_CREATE, audit_logger.ENTITY_EVENT, str(e), entity_name=name)
        flash(f'שגיאה ביצירת האירוע: {str(e)}', 'error')
    
    return redirect(url_for('index'))

@app.route('/events/edit/<int:event_id>', methods=['POST'])
def edit_event(event_id):
    """Edit existing event"""
    vaadot_id = request.form.get('vaadot_id')
    maslul_id = request.form.get('maslul_id')
    name = request.form.get('name', '').strip()
    event_type = request.form.get('event_type')
    expected_requests = request.form.get('expected_requests', '0')
    actual_submissions = request.form.get('actual_submissions', '0')
    call_publication_date = request.form.get('call_publication_date')
    
    # Get manual call deadline parameters
    is_call_deadline_manual = request.form.get('is_call_deadline_manual') == '1'
    manual_call_deadline_date = request.form.get('manual_call_deadline_date')

    if not all([vaadot_id, maslul_id, name, event_type]):
        flash('כל השדות הם שדות חובה', 'error')
        return redirect(url_for('index'))
    
    try:
        expected_requests = int(expected_requests) if expected_requests else 0
        actual_submissions = int(actual_submissions) if actual_submissions else 0
        
        # Validate expected_requests is required and greater than 0
        if expected_requests <= 0:
            flash('מספר הבקשות הצפויות חייב להיות גדול מ-0', 'error')
            return redirect(url_for('index'))
        
        # Validate that committee and route are from the same division
        vaada = None
        for v in db.get_vaadot():
            if v['vaadot_id'] == int(vaadot_id):
                vaada = v
                break
        
        maslul = None
        for m in db.get_maslulim():
            if m['maslul_id'] == int(maslul_id):
                maslul = m
                break
        
        if not vaada or not maslul:
            flash('ועדה או מסלול לא נמצאו במערכת', 'error')
            return redirect(url_for('index'))
        
        if vaada['hativa_id'] != maslul['hativa_id']:
            flash(f'שגיאה: המסלול "{maslul["name"]}" מחטיבת "{maslul["hativa_name"]}" אינו יכול להיות משויך לועדה מחטיבת "{vaada["hativa_name"]}"', 'error')
            return redirect(url_for('index'))
        
        success = db.update_event(event_id, int(vaadot_id), int(maslul_id), name, event_type, expected_requests, actual_submissions, call_publication_date,
                                   is_call_deadline_manual, manual_call_deadline_date)
        if success:
            audit_logger.log_event_updated(event_id, name)
            flash(f'אירוע "{name}" עודכן בהצלחה', 'success')
        else:
            flash('שגיאה בעדכון האירוע', 'error')
    except Exception as e:
        audit_logger.log_error(audit_logger.ACTION_UPDATE, audit_logger.ENTITY_EVENT, str(e), event_id, name)
        flash(f'שגיאה בעדכון האירוע: {str(e)}', 'error')
    
    return redirect(url_for('index'))

@app.route('/events/delete/<int:event_id>', methods=['POST'])
@login_required
def delete_event_route(event_id):
    """Delete event"""
    try:
        # Get current user
        user = auth_manager.get_current_user()
        if not user:
            flash('נדרשת התחברות', 'error')
            return redirect(url_for('login'))
        
        # Check permissions: Users cannot delete events
        if user['role'] == 'user':
            flash('משתמשים רגילים לא יכולים למחוק אירועים', 'error')
            return redirect(url_for('index'))
        
        # Get event name before deletion
        events = db.get_all_events()
        event = next((e for e in events if e['event_id'] == event_id), None)
        
        if not event:
            flash('האירוע לא נמצא במערכת', 'error')
            return redirect(url_for('index'))
        
        # For managers, verify event is in their division
        if user['role'] == 'manager':
            if event.get('maslul_hativa_id') != user['hativa_id']:
                flash('מנהל יכול למחוק רק אירועים בחטיבה שלו', 'error')
                return redirect(url_for('index'))
        
        success = db.delete_event(event_id, user['user_id'])
        if success:
            audit_logger.log_event_deleted(event_id, event['name'])
            flash(f'אירוע "{event["name"]}" הועבר לסל המחזור בהצלחה', 'success')
        else:
            flash('שגיאה במחיקת האירוע', 'error')
    except Exception as e:
        audit_logger.log_error(audit_logger.ACTION_DELETE, audit_logger.ENTITY_EVENT, str(e), event_id)
        flash(f'שגיאה במחיקת האירוע: {str(e)}', 'error')
    
    return redirect(url_for('index'))

@app.route('/hativot/update_color', methods=['POST'])
def update_hativa_color():
    """Update division color"""
    hativa_id = request.form.get('hativa_id')
    color = request.form.get('color')
    
    if not all([hativa_id, color]):
        flash('חטיבה וצבע הם שדות חובה', 'error')
        return redirect(url_for('index'))
    
    try:
        success = db.update_hativa_color(int(hativa_id), color)
        if success:
            flash('צבע החטיבה עודכן בהצלחה', 'success')
        else:
            flash('שגיאה בעדכון צבע החטיבה', 'error')
    except Exception as e:
        flash(f'שגיאה בעדכון צבע החטיבה: {str(e)}', 'error')
    
    return redirect(url_for('index'))







@app.route('/api/maslulim/<int:hativa_id>')
def api_maslulim_by_hativa(hativa_id):
    """API endpoint to get routes by division with enhanced data"""
    try:
        maslulim_list = db.get_maslulim(hativa_id)
        return jsonify({
            'success': True,
            'data': [{
                'maslul_id': m['maslul_id'],
                'name': m['name'],
                'description': m['description'],
                'hativa_name': m['hativa_name']
            } for m in maslulim_list],
            'count': len(maslulim_list)
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'data': [],
            'count': 0
        }), 500

@app.route('/api/vaadot/<int:vaadot_id>/hativa')
def api_vaadot_hativa(vaadot_id):
    """API endpoint to get division ID for a specific committee meeting"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT hativa_id FROM vaadot WHERE vaadot_id = ?', (vaadot_id,))
        result = cursor.fetchone()
        conn.close()
        
        if result:
            return jsonify({
                'success': True,
                'hativa_id': result[0]
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Committee meeting not found'
            }), 404
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/events/<int:event_id>')
def api_event_details(event_id):
    """API endpoint to get event details"""
    try:
        events = db.get_all_events()
        event = next((e for e in events if e['event_id'] == event_id), None)
        
        if event:
            return jsonify({
                'success': True,
                'event': event
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Event not found'
            }), 404
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/maslulim/<int:maslul_id>/details')
def api_maslul_details(maslul_id):
    """API endpoint to get specific route details"""
    try:
        maslulim = db.get_maslulim()
        maslul = next((m for m in maslulim if m['maslul_id'] == maslul_id), None)
        
        if not maslul:
            return jsonify({
                'success': False,
                'error': 'מסלול לא נמצא'
            }), 404
            
        # Get related events count
        events = db.get_all_events()
        events_count = len([e for e in events if e['maslul_id'] == maslul_id])
        
        return jsonify({
            'success': True,
            'data': {
                'maslul_id': maslul['maslul_id'],
                'name': maslul['name'],
                'description': maslul['description'],
                'hativa_name': maslul['hativa_name'],
                'hativa_id': maslul['hativa_id'],
                'events_count': events_count,
                'can_delete': events_count == 0
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/validate_date/<committee_name>/<date_str>')
def api_validate_date(committee_name, date_str):
    """API endpoint to validate committee date"""
    try:
        check_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        # Use auto_scheduler to check if date is available
        can_schedule = auto_scheduler.is_business_day(check_date) and db.is_date_available_for_meeting(check_date)
        reason = "זמין" if can_schedule else "תאריך לא זמין"
        return jsonify({
            'can_schedule': can_schedule,
            'reason': reason
        })
    except ValueError:
        return jsonify({
            'can_schedule': False,
            'reason': 'פורמט תאריך לא תקין'
        })

@app.route('/auto_schedule')
@editor_required
def auto_schedule():
    """Automatic meeting scheduling interface"""
    hativot = db.get_hativot()
    committee_types = db.get_committee_types()
    
    # Get current month by default
    year = request.args.get('year', date.today().year, type=int)
    month = request.args.get('month', date.today().month, type=int)
    
    # Get current user info
    current_user = auth_manager.get_current_user()
    
    return render_template('auto_schedule.html', 
                         hativot=hativot, 
                         committee_types=committee_types,
                         current_year=year,
                         current_month=month,
                         current_user=current_user)

@app.route('/auto_schedule/generate', methods=['POST'])
@editor_required
def generate_auto_schedule():
    try:
        # Get form data
        year = request.form.get('year', type=int)
        month_selection_type = request.form.get('month_selection_type', 'single')
        selected_hativot = request.form.getlist('hativot_ids')
        
        # Get months based on selection type
        months_to_process = []
        if month_selection_type == 'single':
            month = request.form.get('month', type=int)
            months_to_process = [month]
        elif month_selection_type == 'multiple':
            months_to_process = [int(m) for m in request.form.getlist('months')]
        elif month_selection_type == 'range':
            start_month = request.form.get('start_month', type=int)
            end_month = request.form.get('end_month', type=int)
            if start_month <= end_month:
                months_to_process = list(range(start_month, end_month + 1))
            else:
                # Handle year wrap-around (e.g., Nov-Feb)
                months_to_process = list(range(start_month, 13)) + list(range(1, end_month + 1))
        
        if not months_to_process:
            flash('יש לבחור לפחות חודש אחד', 'warning')
            return redirect(url_for('auto_schedule'))
        
        # Determine hativa_id - if only one selected, use it; otherwise None for all
        hativa_id = None
        if len(selected_hativot) == 1:
            hativa_id = int(selected_hativot[0])
        elif len(selected_hativot) > 1:
            # Multiple hativot selected - will process all
            pass
        else:
            flash('יש לבחור לפחות חטיבה אחת', 'warning')
            return redirect(url_for('auto_schedule'))
        
        # Generate schedules for all selected months
        all_suggestions = []
        successful_months = []
        
        from services.auto_schedule_service import ScheduleRequest
        
        for month in months_to_process:
            schedule_request = ScheduleRequest(
                year=year,
                month=month,
                hativa_id=hativa_id,
                auto_approve=False
            )
            
            # Generate schedule using service
            result = auto_schedule_service.generate_schedule(schedule_request)
            
            if result.success and result.suggested_meetings:
                all_suggestions.extend(result.suggested_meetings)
                successful_months.append(month)
                app.logger.info(f"Generated {len(result.suggested_meetings)} suggestions for {year}/{month}")
            else:
                app.logger.warning(f"Failed to generate schedule for {year}/{month}: {result.message}")
        
        if not all_suggestions:
            months_names = [['ינואר', 'פברואר', 'מרץ', 'אפריל', 'מאי', 'יוני',
                            'יולי', 'אוגוסט', 'ספטמבר', 'אוקטובר', 'נובמבר', 'דצמבר'][m-1] 
                           for m in months_to_process]
            flash(f'לא ניתן ליצור תזמון עבור החודשים: {", ".join(months_names)}', 'warning')
            return redirect(url_for('auto_schedule'))
        
        # Store in session for review
        from flask import session
        session['pending_schedule'] = {
            'year': year,
            'months': successful_months,
            'month_selection_type': month_selection_type,
            'suggestions': all_suggestions,
            'selected_hativot': selected_hativot
        }
        
        app.logger.info(f"Stored in session: {len(all_suggestions)} suggestions for {year} months: {successful_months}")
        app.logger.info(f"First suggestion in session: {all_suggestions[0] if all_suggestions else 'None'}")
        
        months_names = [['ינואר', 'פברואר', 'מרץ', 'אפריל', 'מאי', 'יוני',
                        'יולי', 'אוגוסט', 'ספטמבר', 'אוקטובר', 'נובמבר', 'דצמבר'][m-1] 
                       for m in successful_months]
        flash(f'נוצרו {len(all_suggestions)} הצעות ישיבות עבור {", ".join(months_names)} {year}', 'success')
        return redirect(url_for('review_auto_schedule'))
        
    except Exception as e:
        import traceback
        app.logger.error(f'Error generating auto schedule: {str(e)}')
        app.logger.error(f'Traceback: {traceback.format_exc()}')
        flash('שגיאה פנימית ביצירת התזמון. אנא נסה שוב מאוחר יותר.', 'error')
        return redirect(url_for('auto_schedule'))

@app.route('/auto_schedule/review')
@editor_required
def review_auto_schedule():
    """Review generated schedule before approval"""
    from flask import session
    
    pending_schedule = session.get('pending_schedule')
    if not pending_schedule:
        flash('אין לוח זמנים ממתין לאישור', 'warning')
        return redirect(url_for('auto_schedule'))
    
    # Get hativot and committee types for display
    hativot = {h['hativa_id']: h for h in db.get_hativot()}
    committee_types = {ct['committee_type_id']: ct for ct in db.get_committee_types()}
    
    # Enrich suggestions with names
    enriched_suggestions = []
    for suggestion in pending_schedule['suggestions']:
        enriched_suggestions.append({
            **suggestion,
            'hativa_name': hativot.get(suggestion['hativa_id'], {}).get('name', 'לא ידוע'),
            'committee_type_name': suggestion['committee_type']
        })
    
    # Validate schedule constraints for all months
    validation_result = {'valid': True, 'violations': [], 'warnings': []}
    
    # If we have multiple months, validate each one
    months = pending_schedule.get('months', [pending_schedule.get('month')])
    for month in months:
        if month:  # Skip None values
            month_validation = auto_scheduler.validate_schedule_constraints(
                pending_schedule['year'], month
            )
            if not month_validation['valid']:
                validation_result['valid'] = False
                validation_result['violations'].extend(month_validation.get('violations', []))
            validation_result['warnings'].extend(month_validation.get('warnings', []))
    
    # Get current user info
    current_user = auth_manager.get_current_user()
    
    return render_template('review_auto_schedule.html',
                         suggestions=enriched_suggestions,
                         schedule_info=pending_schedule,
                         validation=validation_result,
                         current_user=current_user)

@app.route('/auto_schedule/approve', methods=['POST'])
@editor_required
def approve_auto_schedule():
    """Approve and create selected meetings from the generated schedule"""
    from flask import session
    
    try:
        pending_schedule = session.get('pending_schedule')
        if not pending_schedule:
            flash('אין לוח זמנים ממתין לאישור', 'warning')
            return redirect(url_for('auto_schedule'))
        
        selected_meetings = request.form.getlist('selected_meetings')
        if not selected_meetings:
            flash('יש לבחור לפחות הצעה אחת', 'warning')
            return redirect(url_for('review_auto_schedule'))
        
        # Convert to integers and get selected suggestions
        selected_indices = [int(idx) for idx in selected_meetings]
        suggestions = pending_schedule.get('suggestions', [])
        
        # Filter selected suggestions
        selected_meeting_suggestions = []
        for idx in selected_indices:
            if 0 <= idx < len(suggestions):
                selected_meeting_suggestions.append(suggestions[idx])
        
        if not selected_meeting_suggestions:
            flash('לא נמצאו הצעות תקינות לאישור', 'error')
            return redirect(url_for('review_auto_schedule'))
        
        # Create approval request
        from services.auto_schedule_service import ApprovalRequest
        approval_request = ApprovalRequest(
            suggestions=selected_meeting_suggestions
        )
        
        # Approve meetings using service
        result = auto_schedule_service.approve_meetings(approval_request)
        
        # Clear session
        session.pop('pending_schedule', None)
        
        # Show results
        if result.success_count > 0:
            flash(f'נוצרו בהצלחה {result.success_count} ישיבות', 'success')
        
        if result.failure_count > 0:
            failed_reasons = [f"{m['committee_type']}: {m['reason']}" for m in result.failed_meetings[:3]]
            flash(f'נכשלו {result.failure_count} ישיבות: {"; ".join(failed_reasons)}', 'warning')
        
        return redirect(url_for('index'))
        
    except Exception as e:
        import traceback
        app.logger.error(f'Error approving auto schedule: {str(e)}')
        app.logger.error(f'Traceback: {traceback.format_exc()}')
        flash('שגיאה פנימית באישור התזמון. אנא נסה שוב מאוחר יותר.', 'error')
        return redirect(url_for('review_auto_schedule'))

@app.route('/auto_schedule/validate/<int:year>/<int:month>')
def validate_monthly_schedule(year: int, month: int):
    """API endpoint to validate monthly schedule constraints"""
    try:
        validation_result = auto_schedule_service.get_schedule_validation(year, month)
        return jsonify(validation_result)
    except Exception as e:
        return jsonify({
            'valid': False,
            'violations': [f'שגיאה באימות: {str(e)}'],
            'warnings': [],
            'total_meetings': 0
        })

@app.route('/committee_types')
@editor_required
def committee_types():
    """Manage committee types"""
    # Get hativa_id from query parameters
    hativa_id = request.args.get('hativa_id', type=int)
    
    # Get committee types with statistics (filtered by division if specified)
    response = committee_types_service.get_committee_types_with_statistics(hativa_id)
    
    # Get all hativot for the dropdown
    hativot = db.get_hativot()
    
    if not response.success:
        flash(response.message, 'error')
    
    # Get current user info
    current_user = auth_manager.get_current_user()
    
    return render_template('committee_types.html', 
                         committee_types=response.committee_types,
                         weekly_count=response.statistics['weekly_count'],
                         monthly_count=response.statistics['monthly_count'],
                         active_meetings_count=response.statistics['active_meetings_count'],
                         hativot=hativot,
                         current_user=current_user,
                         selected_hativa_id=hativa_id)

@app.route('/committee_types/add', methods=['POST'])
@editor_required
def add_committee_type():
    """Add new committee type"""
    # Create request object from form data
    committee_type_request = CommitteeTypeRequest(
        hativa_id=request.form.get('hativa_id', type=int),
        name=request.form.get('name', '').strip(),
        scheduled_day=request.form.get('scheduled_day', type=int),
        frequency=request.form.get('frequency', '').strip(),
        week_of_month=request.form.get('week_of_month', type=int) if request.form.get('week_of_month') else None,
        description=request.form.get('description', '').strip()
    )
    
    # Use service to create committee type
    response = committee_types_service.create_committee_type(committee_type_request)
    
    # Log the operation
    if response.success and response.committee_type_id:
        hativot = db.get_hativot()
        hativa = next((h for h in hativot if h['hativa_id'] == committee_type_request.hativa_id), None)
        hativa_name = hativa['name'] if hativa else 'Unknown'
        audit_logger.log_committee_type_created(
            response.committee_type_id,
            committee_type_request.name,
            hativa_name
        )
    elif not response.success:
        audit_logger.log_error(
            audit_logger.ACTION_CREATE,
            audit_logger.ENTITY_COMMITTEE_TYPE,
            response.message,
            entity_name=committee_type_request.name
        )
    
    # Flash appropriate message
    flash(response.message, 'success' if response.success else 'error')
    
    return redirect(url_for('committee_types'))

@app.route('/committee_types/update', methods=['POST'])
@editor_required
def update_committee_type():
    """Update existing committee type"""
    committee_type_id = request.form.get('committee_type_id', type=int)
    
    # Create request object from form data
    committee_type_request = CommitteeTypeRequest(
        hativa_id=request.form.get('hativa_id', type=int),
        name=request.form.get('name', '').strip(),
        scheduled_day=request.form.get('scheduled_day', type=int),
        frequency=request.form.get('frequency', '').strip(),
        week_of_month=request.form.get('week_of_month', type=int) if request.form.get('week_of_month') else None,
        description=request.form.get('description', '').strip()
    )
    
    # Use service to update committee type
    response = committee_types_service.update_committee_type(committee_type_id, committee_type_request)
    
    # Log the operation
    if response.success:
        audit_logger.log_committee_type_updated(committee_type_id, committee_type_request.name)
    else:
        audit_logger.log_error(
            audit_logger.ACTION_UPDATE,
            audit_logger.ENTITY_COMMITTEE_TYPE,
            response.message,
            committee_type_id,
            committee_type_request.name
        )
    
    # Flash appropriate message
    flash(response.message, 'success' if response.success else 'error')
    
    return redirect(url_for('committee_types'))

@app.route('/committee_types/delete', methods=['POST'])
@editor_required
def delete_committee_type():
    """Delete committee type"""
    committee_type_id = request.form.get('committee_type_id', type=int)
    
    # Get committee type name before deletion
    committee_types = db.get_committee_types()
    committee_type = next((ct for ct in committee_types if ct['committee_type_id'] == committee_type_id), None)
    ct_name = committee_type['name'] if committee_type else 'Unknown'
    
    # Use service to delete committee type
    response = committee_types_service.delete_committee_type(committee_type_id)
    
    # Log the operation
    if response.success:
        audit_logger.log_committee_type_deleted(committee_type_id, ct_name)
    else:
        audit_logger.log_error(
            audit_logger.ACTION_DELETE,
            audit_logger.ENTITY_COMMITTEE_TYPE,
            response.message,
            committee_type_id,
            ct_name
        )
    
    # Flash appropriate message
    flash(response.message, 'success' if response.success else 'error')
    
    return redirect(url_for('committee_types'))

# User Management Routes (Admin Only)
@app.route('/admin/users')
@admin_required
def manage_users():
    """User management page"""
    users = db.get_all_users()
    hativot = db.get_hativot()
    current_user = auth_manager.get_current_user()
    
    # Statistics
    total_users = len(users)
    active_users = len([u for u in users if u['is_active']])
    admin_count = len([u for u in users if u['role'] == 'admin'])
    editor_count = len([u for u in users if u['role'] == 'editor'])
    viewer_count = len([u for u in users if u['role'] == 'viewer'])
    
    stats = {
        'total_users': total_users,
        'active_users': active_users,
        'inactive_users': total_users - active_users,
        'admin_count': admin_count,
        'editor_count': editor_count,
        'viewer_count': viewer_count
    }
    
    return render_template('admin/users.html', 
                         users=users, 
                         hativot=hativot, 
                         stats=stats,
                         current_user=current_user)

@app.route('/admin/permissions')
@login_required
def permissions_matrix():
    """Permissions management and matrix view"""
    current_user = auth_manager.get_current_user()
    return render_template('admin/permissions.html', current_user=current_user)

@app.route('/user-guide')
def user_guide():
    """User guide - Single page with all documentation"""
    current_user = auth_manager.get_current_user()
    return render_template('user_guide.html', current_user=current_user)

@app.route('/admin/users/add', methods=['POST'])
@admin_required
def add_user():
    """Add new user with multiple hativot access"""
    try:
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        full_name = request.form.get('full_name', '').strip()
        role = request.form.get('role', 'viewer')
        hativa_ids = request.form.getlist('hativa_ids[]')  # Multiple hativot
        password = request.form.get('password', '').strip()
        
        # Validation
        if not all([username, email, full_name, password]):
            flash('כל השדות הנדרשים חייבים להיות מלאים', 'error')
            return redirect(url_for('manage_users'))
        
        # Validate role
        if role not in ['admin', 'editor', 'viewer']:
            flash('תפקיד לא חוקי', 'error')
            return redirect(url_for('manage_users'))
        
        # Check if username exists
        if db.check_username_exists(username):
            flash('שם המשתמש כבר קיים במערכת', 'error')
            return redirect(url_for('manage_users'))
        
        # Check if email exists
        if db.check_email_exists(email):
            flash('כתובת האימייל כבר קיימת במערכת', 'error')
            return redirect(url_for('manage_users'))
        
        # Convert hativa_ids to integers
        hativa_ids_int = [int(hid) for hid in hativa_ids if hid] if hativa_ids else []
        
        # Hash password
        password_hash = auth_manager.hash_password(password)
        
        # Create user
        user_id = db.create_user(username, email, password_hash, full_name, role, hativa_ids_int)
        
        if user_id:
            audit_logger.log_user_created(user_id, username, role)
            hativot_text = f' עם גישה ל-{len(hativa_ids_int)} חטיבות' if hativa_ids_int else ''
            flash(f'המשתמש {full_name} נוצר בהצלחה{hativot_text}', 'success')
        else:
            flash('שגיאה ביצירת המשתמש', 'error')
            
    except Exception as e:
        flash(f'שגיאה ביצירת המשתמש: {str(e)}', 'error')
    
    return redirect(url_for('manage_users'))

@app.route('/admin/users/update', methods=['POST'])
@admin_required
def update_user():
    """Update user information with multiple hativot access"""
    try:
        user_id = int(request.form.get('user_id'))
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        full_name = request.form.get('full_name', '').strip()
        role = request.form.get('role', 'viewer')
        hativa_ids = request.form.getlist('hativa_ids[]')  # Multiple hativot
        
        # Validation
        if not all([username, email, full_name]):
            flash('כל השדות הנדרשים חייבים להיות מלאים', 'error')
            return redirect(url_for('manage_users'))
        
        # Validate role
        if role not in ['admin', 'editor', 'viewer']:
            flash('תפקיד לא חוקי', 'error')
            return redirect(url_for('manage_users'))
        
        # Check if username exists (excluding current user)
        if db.check_username_exists(username, user_id):
            flash('שם המשתמש כבר קיים במערכת', 'error')
            return redirect(url_for('manage_users'))
        
        # Check if email exists (excluding current user)
        if db.check_email_exists(email, user_id):
            flash('כתובת האימייל כבר קיימת במערכת', 'error')
            return redirect(url_for('manage_users'))
        
        # Convert hativa_ids to integers
        hativa_ids_int = [int(hid) for hid in hativa_ids if hid] if hativa_ids else []
        
        # Update user
        success = db.update_user(user_id, username, email, full_name, role, hativa_ids_int)
        
        if success:
            audit_logger.log_user_updated(user_id, username)
            hativot_text = f' עם גישה ל-{len(hativa_ids_int)} חטיבות' if hativa_ids_int else ''
            flash(f'פרטי המשתמש {full_name} עודכנו בהצלחה{hativot_text}', 'success')
        else:
            flash('שגיאה בעדכון פרטי המשתמש', 'error')
            
    except Exception as e:
        flash(f'שגיאה בעדכון המשתמש: {str(e)}', 'error')
    
    return redirect(url_for('manage_users'))

@app.route('/admin/users/toggle/<int:user_id>', methods=['POST'])
@admin_required
def toggle_user_status(user_id):
    """Toggle user active status"""
    try:
        # Check if trying to deactivate self
        current_user = auth_manager.get_current_user()
        if current_user['user_id'] == user_id:
            flash('לא ניתן להשבית את המשתמש הנוכחי', 'error')
            return redirect(url_for('manage_users'))
        
        success = db.toggle_user_status(user_id)
        
        if success:
            user = db.get_user_by_id(user_id)
            audit_logger.log_user_toggled(user_id, user['username'], user['is_active'])
            status = "הופעל" if user['is_active'] else "הושבת"
            flash(f'המשתמש {user["full_name"]} {status} בהצלחה', 'success')
        else:
            flash('שגיאה בשינוי סטטוס המשתמש', 'error')
            
    except Exception as e:
        flash(f'שגיאה בשינוי סטטוס המשתמש: {str(e)}', 'error')
    
    return redirect(url_for('manage_users'))

@app.route('/admin/users/delete/<int:user_id>', methods=['POST'])
@admin_required
def delete_user(user_id):
    """Delete user (soft delete)"""
    try:
        # Check if trying to delete self
        current_user = auth_manager.get_current_user()
        if current_user['user_id'] == user_id:
            flash('לא ניתן למחוק את המשתמש הנוכחי', 'error')
            return redirect(url_for('manage_users'))
        
        user = db.get_user_by_id(user_id)
        success = db.delete_user(user_id)
        
        if success:
            audit_logger.log_user_deleted(user_id, user['username'])
            flash(f'המשתמש {user["full_name"]} נמחק בהצלחה', 'success')
        else:
            flash('שגיאה במחיקת המשתמש', 'error')
            
    except Exception as e:
        flash(f'שגיאה במחיקת המשתמש: {str(e)}', 'error')
    
    return redirect(url_for('manage_users'))

@app.route('/admin/audit_logs')
@admin_required
def admin_audit_logs():
    """View audit logs (admin only)"""
    try:
        # Get pagination parameters
        page = request.args.get('page', 1, type=int)
        per_page = 50
        offset = (page - 1) * per_page
        
        # Get filter parameters
        username = request.args.get('username', '').strip() or None
        action = request.args.get('action', '').strip() or None
        entity_type = request.args.get('entity_type', '').strip() or None
        status_filter = request.args.get('status', '').strip() or None
        start_date = request.args.get('start_date', '').strip()
        end_date = request.args.get('end_date', '').strip()
        
        # Convert dates
        start_date_obj = None
        end_date_obj = None
        if start_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            except ValueError:
                pass
        if end_date:
            try:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        # Build filters for username search
        user_id_filter = None
        if username:
            # Search for user by username
            all_users = db.get_all_users()
            matching_users = [u for u in all_users if username.lower() in u['username'].lower()]
            if matching_users:
                user_id_filter = matching_users[0]['user_id']
            else:
                # No matching user, use impossible ID
                user_id_filter = -1
        
        # Get logs with filters
        logs = db.get_audit_logs(
            limit=per_page,
            offset=offset,
            user_id=user_id_filter,
            entity_type=entity_type,
            action=action,
            start_date=start_date_obj,
            end_date=end_date_obj
        )
        
        # Filter by status if needed (not in DB method yet)
        if status_filter:
            logs = [log for log in logs if log['status'] == status_filter]
        
        # Get total count
        total_count = db.get_audit_logs_count(
            user_id=user_id_filter,
            entity_type=entity_type,
            action=action,
            start_date=start_date_obj,
            end_date=end_date_obj
        )
        
        total_pages = (total_count + per_page - 1) // per_page
        
        # Get statistics
        stats = db.get_audit_statistics()
        
        # Get current user
        current_user = auth_manager.get_current_user()
        
        return render_template('admin/audit_logs.html',
                             logs=logs,
                             stats=stats,
                             current_page=page,
                             total_pages=total_pages,
                             total_count=total_count,
                             current_user=current_user)
                             
    except Exception as e:
        app.logger.error(f'Error loading audit logs: {str(e)}')
        flash(f'שגיאה בטעינת יומן הביקורת: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/admin/audit_logs/export')
@admin_required
def export_audit_logs():
    """Export audit logs as CSV or Excel"""
    try:
        import csv
        import io
        
        # Get export format (default CSV for backward compatibility)
        export_format = request.args.get('format', 'csv').lower()
        
        # Get filter parameters (same as main route)
        username = request.args.get('username', '').strip() or None
        action = request.args.get('action', '').strip() or None
        entity_type = request.args.get('entity_type', '').strip() or None
        start_date = request.args.get('start_date', '').strip()
        end_date = request.args.get('end_date', '').strip()
        
        # Convert dates
        start_date_obj = None
        end_date_obj = None
        if start_date:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # Get all matching logs
        logs = db.get_audit_logs(
            limit=10000,  # Large limit for export
            offset=0,
            entity_type=entity_type,
            action=action,
            start_date=start_date_obj,
            end_date=end_date_obj
        )
        
        timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if export_format == 'excel':
            # Export as Excel
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Audit Logs"
                
                # Write header with styling
                headers = ['Timestamp', 'Username', 'Action', 'Entity Type', 'Entity Name', 'Details', 'IP Address', 'Status', 'Error Message']
                ws.append(headers)
                
                # Style header row
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Write data
                for log in logs:
                    ws.append([
                        log['timestamp'],
                        log['username'],
                        log['action'],
                        log['entity_type'],
                        log['entity_name'],
                        log['details'],
                        log['ip_address'],
                        log['status'],
                        log['error_message']
                    ])
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Save to BytesIO
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                # Create response
                from flask import make_response
                response = make_response(output.getvalue())
                response.headers['Content-Disposition'] = f'attachment; filename=audit_logs_{timestamp_str}.xlsx'
                response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                
                # Log the export
                audit_logger.log_success(
                    audit_logger.ACTION_EXPORT,
                    'audit_logs',
                    details=f'ייצוא Excel של {len(logs)} רשומות'
                )
                
                return response
                
            except ImportError:
                flash('ספריית openpyxl לא מותקנת. מייצא CSV במקום.', 'warning')
                export_format = 'csv'  # Fall back to CSV
        
        # Export as CSV (default or fallback)
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['Timestamp', 'Username', 'Action', 'Entity Type', 'Entity Name', 'Details', 'IP Address', 'Status', 'Error Message'])
        
        # Write data
        for log in logs:
            writer.writerow([
                log['timestamp'],
                log['username'],
                log['action'],
                log['entity_type'],
                log['entity_name'],
                log['details'],
                log['ip_address'],
                log['status'],
                log['error_message']
            ])
        
        # Create response
        from flask import make_response
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename=audit_logs_{timestamp_str}.csv'
        response.headers['Content-Type'] = 'text/csv'
        
        # Log the export
        audit_logger.log_success(
            audit_logger.ACTION_EXPORT,
            'audit_logs',
            details=f'ייצוא CSV של {len(logs)} רשומות'
        )
        
        return response
        
    except Exception as e:
        app.logger.error(f'Error exporting audit logs: {str(e)}')
        flash(f'שגיאה בייצוא יומן הביקורת: {str(e)}', 'error')
        return redirect(url_for('admin_audit_logs'))

@app.route('/admin/users/change_password', methods=['POST'])
@admin_required
def change_user_password():
    """Change user password"""
    try:
        user_id = int(request.form.get('user_id'))
        new_password = request.form.get('new_password', '').strip()
        
        if not new_password:
            flash('נדרשת סיסמה חדשה', 'error')
            return redirect(url_for('manage_users'))
        
        if len(new_password) < 6:
            flash('הסיסמה חייבת להכיל לפחות 6 תווים', 'error')
            return redirect(url_for('manage_users'))
        
        # Hash new password
        new_password_hash = auth_manager.hash_password(new_password)
        
        # Update password
        success = db.change_user_password(user_id, new_password_hash)
        
        if success:
            user = db.get_user_by_id(user_id)
            audit_logger.log_user_password_changed(user_id, user['username'], by_admin=True)
            flash(f'הסיסמה של {user["full_name"]} שונתה בהצלחה', 'success')
        else:
            flash('שגיאה בשינוי הסיסמה', 'error')
            
    except Exception as e:
        flash(f'שגיאה בשינוי הסיסמה: {str(e)}', 'error')
    
    return redirect(url_for('manage_users'))

# Active Directory Settings Routes
@app.route('/admin/ad_settings')
@admin_required
def ad_settings():
    """Active Directory settings page"""
    try:
        # Get all AD settings
        ad_config = {
            'enabled': db.get_system_setting('ad_enabled') == '1',
            'auth_method': 'oauth',  # Always OAuth for Azure AD
            
            # Azure AD OAuth Settings - from .env file
            'azure_tenant_id': os.getenv('AZURE_TENANT_ID', ''),
            'azure_client_id': os.getenv('AZURE_CLIENT_ID', ''),
            'azure_client_secret': os.getenv('AZURE_CLIENT_SECRET', ''),
            'azure_redirect_uri': os.getenv('AZURE_REDIRECT_URI', ''),
            
            # Common Settings - from database
            'admin_group': db.get_system_setting('ad_admin_group') or '',
            'manager_group': db.get_system_setting('ad_manager_group') or '',
            'auto_create_users': db.get_system_setting('ad_auto_create_users') == '1',
            'default_hativa_id': db.get_system_setting('ad_default_hativa_id') or '',
            'sync_on_login': db.get_system_setting('ad_sync_on_login') == '1'
        }
        
        # Get hativot for default division selection
        hativot = db.get_hativot()
        
        # Get AD users count
        ad_users = db.get_ad_users()
        local_users = db.get_local_users()
        
        current_user = auth_manager.get_current_user()
        
        return render_template('admin/ad_settings.html',
                             ad_config=ad_config,
                             hativot=hativot,
                             ad_users_count=len(ad_users),
                             local_users_count=len(local_users),
                             current_user=current_user)
    except Exception as e:
        app.logger.error(f'Error loading AD settings: {str(e)}')
        flash(f'שגיאה בטעינת הגדרות Active Directory: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/admin/ad_settings/update', methods=['POST'])
@admin_required
def update_ad_settings():
    """Update Active Directory settings"""
    try:
        user_id = session['user_id']
        
        # Get form data - only save general settings (Azure credentials are in .env)
        settings = {
            'ad_enabled': '1' if request.form.get('enabled') == 'on' else '0',
            'ad_auth_method': 'oauth',  # Always OAuth for Azure AD
            
            # Common Settings
            'ad_admin_group': request.form.get('admin_group', '').strip(),
            'ad_manager_group': request.form.get('manager_group', '').strip(),
            'ad_auto_create_users': '1' if request.form.get('auto_create_users') == 'on' else '0',
            'ad_default_hativa_id': request.form.get('default_hativa_id', '').strip(),
            'ad_sync_on_login': '1' if request.form.get('sync_on_login') == 'on' else '0'
        }
        
        # Note: Azure AD credentials (tenant_id, client_id, client_secret, redirect_uri) 
        # are now managed in .env file for better security
        
        # Update all settings
        for key, value in settings.items():
            db.update_system_setting(key, value, user_id)
        
        # Reload AD service with new settings
        ad_service.reload_settings()
        
        audit_logger.log_success(
            audit_logger.ACTION_UPDATE,
            'ad_settings',
            details='עדכון הגדרות Active Directory'
        )
        
        flash('הגדרות Active Directory עודכנו בהצלחה', 'success')
        
    except Exception as e:
        audit_logger.log_error(
            audit_logger.ACTION_UPDATE,
            'ad_settings',
            str(e),
            details='עדכון הגדרות Active Directory'
        )
        flash(f'שגיאה בעדכון הגדרות AD: {str(e)}', 'error')
    
    return redirect(url_for('ad_settings'))

@app.route('/admin/ad_settings/test', methods=['POST'])
@admin_required
def test_ad_connection():
    """Test Active Directory connection"""
    try:
        # Reload settings first
        ad_service.reload_settings()
        
        # Test connection
        success, message = ad_service.test_connection()
        
        if success:
            audit_logger.log_success(
                audit_logger.ACTION_TEST,
                'ad_connection',
                details='בדיקת חיבור AD הצליחה'
            )
            return jsonify({'success': True, 'message': message})
        else:
            audit_logger.log_error(
                audit_logger.ACTION_TEST,
                'ad_connection',
                message,
                details='בדיקת חיבור AD נכשלה'
            )
            return jsonify({'success': False, 'message': message})
            
    except Exception as e:
        app.logger.error(f'Error testing AD connection: {str(e)}')
        return jsonify({'success': False, 'message': f'שגיאה: {str(e)}'})

@app.route('/admin/ad_settings/test_azure', methods=['POST'])
@admin_required
def test_azure_connection():
    """Test Azure AD OAuth configuration"""
    try:
        # Reload settings first
        ad_service.reload_settings()
        
        # Test Azure AD configuration
        success, message = ad_service.test_azure_connection()
        
        if success:
            audit_logger.log_success(
                audit_logger.ACTION_TEST,
                'azure_ad_config',
                details='בדיקת הגדרות Azure AD הצליחה'
            )
            return jsonify({'success': True, 'message': message})
        else:
            audit_logger.log_error(
                audit_logger.ACTION_TEST,
                'azure_ad_config',
                message,
                details='בדיקת הגדרות Azure AD נכשלה'
            )
            return jsonify({'success': False, 'message': message})
            
    except Exception as e:
        app.logger.error(f'Error testing Azure AD configuration: {str(e)}')
        return jsonify({'success': False, 'message': f'שגיאה: {str(e)}'})

@app.route('/admin/ad_settings/search_users', methods=['POST'])
@admin_required
def search_ad_users():
    """Search for users in Active Directory"""
    try:
        search_term = request.form.get('search_term', '').strip()
        
        if not search_term:
            return jsonify({'success': False, 'message': 'נדרש מונח חיפוש'})
        
        # Search AD
        users = ad_service.search_users(search_term, limit=20)
        
        return jsonify({'success': True, 'users': users, 'count': len(users)})
        
    except Exception as e:
        app.logger.error(f'Error searching AD users: {str(e)}')
        return jsonify({'success': False, 'message': f'שגיאה בחיפוש: {str(e)}'})

@app.route('/admin/ad_settings/sync_user', methods=['POST'])
@admin_required
def sync_ad_user():
    """Manually sync a user from AD"""
    try:
        username = request.form.get('username', '').strip()
        role = request.form.get('role', 'user')
        hativa_id = request.form.get('hativa_id')
        
        if not username:
            return jsonify({'success': False, 'message': 'נדרש שם משתמש'})
        
        # Get user from AD
        _, user_info, _ = ad_service.authenticate(username, '')
        
        if not user_info:
            # Try to search for user
            users = ad_service.search_users(username, limit=1)
            if users:
                user_info = users[0]
            else:
                return jsonify({'success': False, 'message': 'משתמש לא נמצא ב-AD'})
        
        # Sync to local DB
        hativa_id_int = int(hativa_id) if hativa_id else None
        user_id = ad_service.sync_user_to_local(user_info, role, hativa_id_int)
        
        if user_id:
            audit_logger.log_user_created(user_id, username, role)
            return jsonify({'success': True, 'message': f'משתמש {username} סונכרן בהצלחה'})
        else:
            return jsonify({'success': False, 'message': 'שגיאה בסנכרון המשתמש'})
            
    except Exception as e:
        app.logger.error(f'Error syncing AD user: {str(e)}')
        return jsonify({'success': False, 'message': f'שגיאה: {str(e)}'})

# Drag & Drop API endpoints
@app.route('/api/move_committee', methods=['POST'])
@login_required
def move_committee():
    """Move committee meeting to a different date"""
    try:
        data = request.get_json()
        vaada_id = data.get('vaada_id')
        new_date = data.get('new_date')
        
        if not vaada_id or not new_date:
            return jsonify({'success': False, 'message': 'נתונים חסרים'}), 400
        
        # Get current user
        user = auth_manager.get_current_user()
        if not user:
            return jsonify({'success': False, 'message': 'נדרשת התחברות'}), 401
        
        # Get committee details to check permissions
        vaada = db.get_vaada_by_id(vaada_id)
        if not vaada:
            return jsonify({'success': False, 'message': 'ועדה לא נמצאה'}), 404
        
        # Check permissions: Only managers and admins can move committees
        if user['role'] == 'user':
            return jsonify({'success': False, 'message': 'רק מנהלים ומנהלי מערכת יכולים להזיז ועדות'}), 403
        
        # Manager can only move committees in their division
        if user['role'] == 'manager':
            if vaada['hativa_id'] != user['hativa_id']:
                return jsonify({'success': False, 'message': 'מנהל יכול להזיז רק ועדות בחטיבה שלו'}), 403
        
        # Admin can move any committee (no additional checks needed)
        
        # Validate date format and convert to date object
        try:
            new_date_obj = datetime.strptime(new_date, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'success': False, 'message': 'פורמט תאריך לא תקין'}), 400
        
        # Ensure target date is an allowed business day
        if not db.is_work_day(new_date_obj):
            return jsonify({'success': False, 'message': 'לא ניתן להעביר ועדה ליום שאינו יום עסקים'}), 400

        # Check if target date is available (one meeting per day constraint)
        # But allow moving the same committee to a different date
        existing_meetings = db.get_vaada_by_date(new_date_obj)
        if existing_meetings and any(meeting['vaadot_id'] != vaada_id for meeting in existing_meetings):
            return jsonify({'success': False, 'message': 'התאריך תפוס - יש כבר ועדה ביום זה'}), 400
        
        # Store old date and committee name for logging
        old_date = vaada['vaada_date'] if vaada else 'Unknown'
        committee_name = vaada['committee_name'] if vaada else 'Unknown'
        
        # Update committee meeting date
        try:
            success = db.update_vaada_date(vaada_id, new_date_obj)
        except ValueError as ve:
            audit_logger.log_error(
                audit_logger.ACTION_MOVE,
                audit_logger.ENTITY_VAADA,
                str(ve),
                entity_id=vaada_id,
                entity_name=committee_name,
                details=f'נסיון העברה מ-{old_date} ל-{new_date}'
            )
            return jsonify({'success': False, 'message': str(ve)}), 400
        
        if success:
            # Log successful move
            audit_logger.log_vaada_moved(vaada_id, committee_name, old_date, new_date)
            return jsonify({'success': True, 'message': 'הועדה הועברה בהצלחה'})
        else:
            audit_logger.log_error(
                audit_logger.ACTION_MOVE,
                audit_logger.ENTITY_VAADA,
                'שגיאה בעדכון מסד נתונים',
                entity_id=vaada_id,
                entity_name=committee_name
            )
            return jsonify({'success': False, 'message': 'שגיאה בהעברת הועדה'}), 500
            
    except Exception as e:
        app.logger.error(f"Error moving committee: {str(e)}")
        audit_logger.log_error(
            audit_logger.ACTION_MOVE,
            audit_logger.ENTITY_VAADA,
            str(e),
            entity_id=vaada_id if 'vaada_id' in locals() else None
        )
        return jsonify({'success': False, 'message': f'שגיאה: {str(e)}'}), 500

@app.route('/api/move_event', methods=['POST'])
@login_required
@editing_permission_required
def move_event():
    """Move event to a different committee meeting"""
    try:
        data = request.get_json()
        event_id = data.get('event_id')
        target_vaada_id = data.get('target_vaada_id')
        
        if not event_id or not target_vaada_id:
            return jsonify({'success': False, 'message': 'נתונים חסרים'}), 400
        
        # Get event and target committee details for validation
        app.logger.info(f"Looking for event_id: {event_id}, target_vaada_id: {target_vaada_id}")
        event = db.get_event_by_id(event_id)
        target_committee = db.get_vaada_by_id(target_vaada_id)
        
        app.logger.info(f"Found event: {event}")
        app.logger.info(f"Found target_committee: {target_committee}")
        
        if not event or not target_committee:
            app.logger.error(f"Event or committee not found. Event: {event}, Committee: {target_committee}")
            return jsonify({'success': False, 'message': 'אירוע או ועדה לא נמצאו'}), 404
        
        # Validate that event's route belongs to target committee's division
        route = db.get_maslul_by_id(event['maslul_id'])
        if route['hativa_id'] != target_committee['hativa_id']:
            audit_logger.log_error(
                audit_logger.ACTION_MOVE,
                audit_logger.ENTITY_EVENT,
                'ניסיון להעביר אירוע לועדה מחטיבה אחרת',
                entity_id=event_id,
                entity_name=event.get('name', 'Unknown')
            )
            return jsonify({'success': False, 'message': 'לא ניתן להעביר אירוע לועדה מחטיבה אחרת'}), 400
        
        # Get source committee name for logging
        source_vaada = db.get_vaada_by_id(event.get('vaadot_id'))
        source_committee_name = source_vaada['committee_name'] if source_vaada else 'Unknown'
        
        # Update event's committee meeting (includes max requests validation)
        success = db.update_event_vaada(event_id, target_vaada_id)
        
        if success:
            # Log successful move
            audit_logger.log_event_moved(
                event_id,
                event.get('name', 'Unknown'),
                source_committee_name,
                target_committee['committee_name']
            )
            return jsonify({'success': True, 'message': 'האירוע הועבר בהצלחה'})
        else:
            audit_logger.log_error(
                audit_logger.ACTION_MOVE,
                audit_logger.ENTITY_EVENT,
                'שגיאה בעדכון מסד נתונים',
                entity_id=event_id,
                entity_name=event.get('name', 'Unknown')
            )
            return jsonify({'success': False, 'message': 'שגיאה בהעברת האירוע'}), 500
    
    except ValueError as e:
        # Handle constraint violations
        app.logger.warning(f"Event move blocked by constraint: {str(e)}")
        audit_logger.log_error(
            audit_logger.ACTION_MOVE,
            audit_logger.ENTITY_EVENT,
            str(e),
            entity_id=event_id if 'event_id' in locals() else None
        )
        return jsonify({'success': False, 'message': str(e)}), 400
            
    except Exception as e:
        app.logger.error(f"Error moving event: {str(e)}")
        audit_logger.log_error(
            audit_logger.ACTION_MOVE,
            audit_logger.ENTITY_EVENT,
            str(e),
            entity_id=event_id if 'event_id' in locals() else None
        )
        return jsonify({'success': False, 'message': f'שגיאה: {str(e)}'}), 500

@app.route('/api/events_by_committee')
@login_required
def get_events_by_committee():
    """API endpoint to get events grouped by committee meetings"""
    try:
        # Get all events
        events = db.get_all_events()
        include_empty = request.args.get('include_empty') in ('1', 'true', 'True')

        # Group events by committee meeting (vaadot_id)
        events_by_committee = {}
        for event in events:
            vaadot_id = event.get('vaadot_id')
            if vaadot_id:
                if vaadot_id not in events_by_committee:
                    events_by_committee[vaadot_id] = {
                        'committee_info': {
                            'vaadot_id': vaadot_id,
                            'committee_name': event.get('committee_name', ''),
                            'hativa_name': event.get('hativa_name', ''),
                            'vaada_date': event.get('vaada_date', ''),
                            'committee_type': event.get('committee_type_name', '')
                        },
                        'events': [],
                        'summary': {
                            'total_events': 0,
                            'total_expected_requests': 0,
                            'total_actual_submissions': 0,
                            'event_types': set()
                        }
                    }
                events_by_committee[vaadot_id]['events'].append(event)

                # Update summary
                summary = events_by_committee[vaadot_id]['summary']
                summary['total_events'] += 1
                summary['total_expected_requests'] += event.get('expected_requests', 0) or 0
                summary['total_actual_submissions'] += event.get('actual_submissions', 0) or 0
                if event.get('event_type'):
                    summary['event_types'].add(event.get('event_type'))

        # Optionally include committees without events
        if include_empty:
            committees = db.get_vaadot()
            for c in committees:
                vid = c.get('vaadot_id')
                if vid not in events_by_committee:
                    events_by_committee[vid] = {
                        'committee_info': {
                            'vaadot_id': vid,
                            'committee_name': c.get('committee_name', ''),
                            'hativa_name': c.get('hativa_name', ''),
                            'vaada_date': c.get('vaada_date', ''),
                            'committee_type': ''
                        },
                        'events': [],
                        'summary': {
                            'total_events': 0,
                            'total_expected_requests': 0,
                            'total_actual_submissions': 0,
                            'event_types': set()
                        }
                    }

        # Convert to list and format for JSON response
        result = []
        for vaadot_id, data in events_by_committee.items():
            # Convert set to list for JSON serialization
            data['summary']['event_types'] = list(data['summary']['event_types'])
            result.append(data)

        # Sort by committee date (newest first)
        result.sort(key=lambda x: x['committee_info']['vaada_date'] or '', reverse=True)

        return jsonify({
            'success': True,
            'events_by_committee': result,
            'total_committees': len(result)
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Recycle Bin Routes
@app.route('/recycle_bin')
@editor_required
def recycle_bin():
    """Recycle bin view for deleted items"""
    try:
        current_user = auth_manager.get_current_user()
        
        # Managers can only see their division's deleted items
        hativa_id = None
        if current_user['role'] == 'manager':
            hativa_id = current_user['hativa_id']
        
        deleted_vaadot = db.get_deleted_vaadot(hativa_id)
        deleted_events = db.get_deleted_events(hativa_id)
        
        return render_template('recycle_bin.html',
                             deleted_vaadot=deleted_vaadot,
                             deleted_events=deleted_events,
                             current_user=current_user)
    except Exception as e:
        flash(f'שגיאה בטעינת סל המחזור: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/api/recycle_bin/restore_vaada/<int:vaadot_id>', methods=['POST'])
@editor_required
def restore_vaada_route(vaadot_id):
    """Restore a deleted committee meeting"""
    try:
        user = auth_manager.get_current_user()
        if not user:
            return jsonify({'success': False, 'message': 'נדרשת התחברות'}), 401
        
        # Users cannot restore
        if user['role'] == 'user':
            return jsonify({'success': False, 'message': 'משתמשים רגילים לא יכולים לשחזר פריטים'}), 403
        
        # Get the vaada to check permissions for managers
        deleted_vaadot = db.get_deleted_vaadot()
        vaada = next((v for v in deleted_vaadot if v['vaadot_id'] == vaadot_id), None)
        
        if not vaada:
            return jsonify({'success': False, 'message': 'ועדה לא נמצאה בסל המחזור'}), 404
        
        # Check manager permissions
        if user['role'] == 'manager' and vaada['hativa_id'] != user['hativa_id']:
            return jsonify({'success': False, 'message': 'מנהל יכול לשחזר רק ועדות מהחטיבה שלו'}), 403
        
        success = db.restore_vaada(vaadot_id)
        if success:
            audit_logger.log_success(
                audit_logger.ACTION_CREATE,
                audit_logger.ENTITY_VAADA,
                details=f'שחזור ועדה מסל המחזור: {vaada["committee_name"]}'
            )
            return jsonify({'success': True, 'message': 'הועדה שוחזרה בהצלחה'})
        else:
            return jsonify({'success': False, 'message': 'שגיאה בשחזור הועדה'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/recycle_bin/restore_event/<int:event_id>', methods=['POST'])
@editor_required
def restore_event_route(event_id):
    """Restore a deleted event"""
    try:
        user = auth_manager.get_current_user()
        if not user:
            return jsonify({'success': False, 'message': 'נדרשת התחברות'}), 401
        
        # Users cannot restore
        if user['role'] == 'user':
            return jsonify({'success': False, 'message': 'משתמשים רגילים לא יכולים לשחזר פריטים'}), 403
        
        # Get the event to check permissions for managers
        deleted_events = db.get_deleted_events()
        event = next((e for e in deleted_events if e['event_id'] == event_id), None)
        
        if not event:
            return jsonify({'success': False, 'message': 'אירוע לא נמצא בסל המחזור'}), 404
        
        # Check manager permissions
        if user['role'] == 'manager' and event['maslul_hativa_id'] != user['hativa_id']:
            return jsonify({'success': False, 'message': 'מנהל יכול לשחזר רק אירועים מהחטיבה שלו'}), 403
        
        success = db.restore_event(event_id)
        if success:
            audit_logger.log_success(
                audit_logger.ACTION_CREATE,
                audit_logger.ENTITY_EVENT,
                details=f'שחזור אירוע מסל המחזור: {event["name"]}'
            )
            return jsonify({'success': True, 'message': 'האירוע שוחזר בהצלחה'})
        else:
            return jsonify({'success': False, 'message': 'שגיאה בשחזור האירוע'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/recycle_bin/permanent_delete_vaada/<int:vaadot_id>', methods=['POST'])
@editor_required
def permanent_delete_vaada_route(vaadot_id):
    """Permanently delete a committee meeting"""
    try:
        user = auth_manager.get_current_user()
        if not user:
            return jsonify({'success': False, 'message': 'נדרשת התחברות'}), 401
        
        # Only admins can permanently delete
        if user['role'] != 'admin':
            return jsonify({'success': False, 'message': 'רק מנהלי מערכת יכולים למחוק לצמיתות'}), 403
        
        # Get the vaada info for logging
        deleted_vaadot = db.get_deleted_vaadot()
        vaada = next((v for v in deleted_vaadot if v['vaadot_id'] == vaadot_id), None)
        
        if not vaada:
            return jsonify({'success': False, 'message': 'ועדה לא נמצאה בסל המחזור'}), 404
        
        success = db.permanently_delete_vaada(vaadot_id)
        if success:
            audit_logger.log_success(
                audit_logger.ACTION_DELETE,
                audit_logger.ENTITY_VAADA,
                details=f'מחיקה לצמיתות: {vaada["committee_name"]}'
            )
            return jsonify({'success': True, 'message': 'הועדה נמחקה לצמיתות'})
        else:
            return jsonify({'success': False, 'message': 'שגיאה במחיקה לצמיתות'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/recycle_bin/permanent_delete_event/<int:event_id>', methods=['POST'])
@editor_required
def permanent_delete_event_route(event_id):
    """Permanently delete an event"""
    try:
        user = auth_manager.get_current_user()
        if not user:
            return jsonify({'success': False, 'message': 'נדרשת התחברות'}), 401
        
        # Only admins can permanently delete
        if user['role'] != 'admin':
            return jsonify({'success': False, 'message': 'רק מנהלי מערכת יכולים למחוק לצמיתות'}), 403
        
        # Get the event info for logging
        deleted_events = db.get_deleted_events()
        event = next((e for e in deleted_events if e['event_id'] == event_id), None)
        
        if not event:
            return jsonify({'success': False, 'message': 'אירוע לא נמצא בסל המחזור'}), 404
        
        success = db.permanently_delete_event(event_id)
        if success:
            audit_logger.log_success(
                audit_logger.ACTION_DELETE,
                audit_logger.ENTITY_EVENT,
                details=f'מחיקה לצמיתות: {event["name"]}'
            )
            return jsonify({'success': True, 'message': 'האירוע נמחק לצמיתות'})
        else:
            return jsonify({'success': False, 'message': 'שגיאה במחיקה לצמיתות'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/recycle_bin/empty', methods=['POST'])
@editor_required
def empty_recycle_bin_route():
    """Empty the recycle bin (permanently delete all items)"""
    try:
        user = auth_manager.get_current_user()
        if not user:
            return jsonify({'success': False, 'message': 'נדרשת התחברות'}), 401
        
        # Only admins can empty recycle bin
        if user['role'] != 'admin':
            return jsonify({'success': False, 'message': 'רק מנהלי מערכת יכולים לרוקן את סל המחזור'}), 403
        
        vaadot_deleted, events_deleted = db.empty_recycle_bin()
        
        audit_logger.log_success(
            audit_logger.ACTION_DELETE,
            'recycle_bin',
            details=f'ריקון סל מחזור: {vaadot_deleted} ועדות, {events_deleted} אירועים'
        )
        
        return jsonify({
            'success': True, 
            'message': f'סל המחזור רוקן בהצלחה',
            'vaadot_deleted': vaadot_deleted,
            'events_deleted': events_deleted
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/events_table')
@login_required
def events_table():
    """Events table view with advanced filtering"""
    try:
        # Get all events with extended information
        events = db.get_all_events()

        # Normalize date fields for consistent formatting in the template
        from datetime import datetime
        for event in events:
            created_at = event.get('created_at')
            if isinstance(created_at, str):
                for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
                    try:
                        event['created_at'] = datetime.strptime(created_at, fmt)
                        break
                    except ValueError:
                        continue

            for date_field in ['call_deadline_date', 'intake_deadline_date', 'review_deadline_date', 'response_deadline_date', 'vaada_date']:
                value = event.get(date_field)
                if isinstance(value, str):
                    for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S'):
                        try:
                            event[date_field] = datetime.strptime(value, fmt).date()
                            break
                        except ValueError:
                            continue

        # Get filter options
        hativot = db.get_hativot()
        maslulim = db.get_maslulim()
        committee_types = db.get_committee_types()

        # Get unique event types
        event_types = list(set([event.get('event_type', '') for event in events if event.get('event_type')]))

        # Get current user info
        current_user = auth_manager.get_current_user()

        return render_template('events_table.html', 
                             events=events,
                             hativot=hativot,
                             maslulim=maslulim,
                             committee_types=committee_types,
                             event_types=event_types,
                             current_user=current_user)
    except Exception as e:
        flash(f'שגיאה בטעינת נתוני האירועים: {str(e)}', 'error')
        return redirect(url_for('index'))

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5001))
    debug = os.environ.get('FLASK_ENV') != 'production'
    app.run(debug=debug, host='0.0.0.0', port=port)
