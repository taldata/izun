from flask import Blueprint, render_template, request, flash, redirect, url_for, session, jsonify, current_app
from services_init import db, auth_manager, audit_logger, ad_service
from auth import admin_required
import os
from datetime import datetime

settings_bp = Blueprint('settings', __name__)

@settings_bp.route('/admin/audit_logs')
@admin_required
def admin_audit_logs():
    """View audit logs (admin only)"""
    try:
        # Get pagination parameters
        page = request.args.get('page', 1, type=int)
        per_page = 50
        offset = (page - 1) * per_page
        
        # Get filter parameters
        username = request.args.get('username', '').strip() or None
        action = request.args.get('action', '').strip() or None
        entity_type = request.args.get('entity_type', '').strip() or None
        status_filter = request.args.get('status', '').strip() or None
        search_text = request.args.get('search_text', '').strip() or None
        start_date = request.args.get('start_date', '').strip()
        end_date = request.args.get('end_date', '').strip()
        
        # Convert dates
        start_date_obj = None
        end_date_obj = None
        if start_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            except ValueError:
                pass
        if end_date:
            try:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        # Build filters for username search
        user_id_filter = None
        if username:
            # Search for user by username
            all_users = db.get_all_users()
            matching_users = [u for u in all_users if username.lower() in u['username'].lower()]
            if matching_users:
                user_id_filter = matching_users[0]['user_id']
            else:
                # No matching user, use impossible ID
                user_id_filter = -1
        
        # Get logs with filters
        logs = db.get_audit_logs(
            limit=per_page,
            offset=offset,
            user_id=user_id_filter,
            entity_type=entity_type,
            action=action,
            search_text=search_text,
            start_date=start_date_obj,
            end_date=end_date_obj
        )
        
        # Filter by status if needed (not in DB method yet)
        if status_filter:
            logs = [log for log in logs if log['status'] == status_filter]
        
        # Get total count
        total_count = db.get_audit_logs_count(
            user_id=user_id_filter,
            entity_type=entity_type,
            action=action,
            search_text=search_text,
            start_date=start_date_obj,
            end_date=end_date_obj
        )
        
        total_pages = (total_count + per_page - 1) // per_page
        
        # Get statistics
        stats = db.get_audit_statistics()
        
        # Get current user
        current_user = auth_manager.get_current_user()
        
        return render_template('admin/audit_logs.html',
                             logs=logs,
                             stats=stats,
                             current_page=page,
                             total_pages=total_pages,
                             total_count=total_count,
                             current_user=current_user)
                             
    except Exception as e:
        current_app.logger.error(f'Error loading audit logs: {str(e)}')
        flash(f'שגיאה בטעינת יומן הביקורת: {str(e)}', 'error')
        return redirect(url_for('main.index'))

@settings_bp.route('/admin/audit_logs/export')
@admin_required
def export_audit_logs():
    """Export audit logs as CSV or Excel"""
    try:
        import csv
        import io
        
        # Get export format (default CSV for backward compatibility)
        export_format = request.args.get('format', 'csv').lower()
        
        # Get filter parameters (same as main route)
        username = request.args.get('username', '').strip() or None
        action = request.args.get('action', '').strip() or None
        entity_type = request.args.get('entity_type', '').strip() or None
        search_text = request.args.get('search_text', '').strip() or None
        start_date = request.args.get('start_date', '').strip()
        end_date = request.args.get('end_date', '').strip()
        
        # Convert dates
        start_date_obj = None
        end_date_obj = None
        if start_date:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # Get all matching logs
        logs = db.get_audit_logs(
            limit=10000,  # Large limit for export
            offset=0,
            entity_type=entity_type,
            action=action,
            search_text=search_text,
            start_date=start_date_obj,
            end_date=end_date_obj
        )
        
        timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if export_format == 'excel':
            # Export as Excel
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Audit Logs"
                
                # Write header with styling
                headers = ['Timestamp', 'Username', 'Action', 'Entity Type', 'Entity Name', 'Details', 'IP Address', 'Status', 'Error Message']
                ws.append(headers)
                
                # Style header row
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Write data
                for log in logs:
                    ws.append([
                        log['timestamp'],
                        log['username'],
                        log['action'],
                        log['entity_type'],
                        log['entity_name'],
                        log['details'],
                        log['ip_address'],
                        log['status'],
                        log['error_message']
                    ])
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Save to BytesIO
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                # Create response
                from flask import make_response
                response = make_response(output.getvalue())
                response.headers['Content-Disposition'] = f'attachment; filename=audit_logs_{timestamp_str}.xlsx'
                response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                
                # Log the export
                audit_logger.log_success(
                    audit_logger.ACTION_EXPORT,
                    'audit_logs',
                    details=f'ייצוא Excel של {len(logs)} רשומות'
                )
                
                return response
                
            except ImportError:
                flash('ספריית openpyxl לא מותקנת. מייצא CSV במקום.', 'warning')
                export_format = 'csv'  # Fall back to CSV
        
        # Export as CSV (default or fallback)
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['Timestamp', 'Username', 'Action', 'Entity Type', 'Entity Name', 'Details', 'IP Address', 'Status', 'Error Message'])
        
        # Write data
        for log in logs:
            writer.writerow([
                log['timestamp'],
                log['username'],
                log['action'],
                log['entity_type'],
                log['entity_name'],
                log['details'],
                log['ip_address'],
                log['status'],
                log['error_message']
            ])
        
        # Create response
        from flask import make_response
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename=audit_logs_{timestamp_str}.csv'
        response.headers['Content-Type'] = 'text/csv'
        
        # Log the export
        audit_logger.log_success(
            audit_logger.ACTION_EXPORT,
            'audit_logs',
            details=f'ייצוא CSV של {len(logs)} רשומות'
        )
        
        return response
        
    except Exception as e:
        current_app.logger.error(f'Error exporting audit logs: {str(e)}')
        flash(f'שגיאה בייצוא יומן הביקורת: {str(e)}', 'error')
        return redirect(url_for('settings.admin_audit_logs'))

# Active Directory Settings Routes
@settings_bp.route('/admin/ad_settings')
@admin_required
def ad_settings():
    """Active Directory settings page"""
    try:
        # Get all AD settings
        ad_config = {
            'enabled': db.get_system_setting('ad_enabled') == '1',
            'auth_method': 'oauth',  # Always OAuth for Azure AD
            
            # Azure AD OAuth Settings - from .env file
            'azure_tenant_id': os.getenv('AZURE_TENANT_ID', ''),
            'azure_client_id': os.getenv('AZURE_CLIENT_ID', ''),
            'azure_client_secret': os.getenv('AZURE_CLIENT_SECRET', ''),
            'azure_redirect_uri': os.getenv('AZURE_REDIRECT_URI', ''),
            
            # Common Settings - from database
            'admin_group': db.get_system_setting('ad_admin_group') or '',
            'manager_group': db.get_system_setting('ad_manager_group') or '',
            'auto_create_users': db.get_system_setting('ad_auto_create_users') == '1',
            'default_hativa_id': db.get_system_setting('ad_default_hativa_id') or '',
            'sync_on_login': db.get_system_setting('ad_sync_on_login') == '1'
        }
        
        # Get hativot for default division selection
        hativot = db.get_hativot()
        
        # Get AD users count
        ad_users = db.get_ad_users()
        
        current_user = auth_manager.get_current_user()
        
        return render_template('admin/ad_settings.html',
                             ad_config=ad_config,
                             hativot=hativot,
                             ad_users_count=len(ad_users),
                             current_user=current_user)
    except Exception as e:
        current_app.logger.error(f'Error loading AD settings: {str(e)}')
        flash(f'שגיאה בטעינת הגדרות Active Directory: {str(e)}', 'error')
        return redirect(url_for('main.index'))

@settings_bp.route('/admin/ad_settings/update', methods=['POST'])
@admin_required
def update_ad_settings():
    """Update Active Directory settings"""
    try:
        user_id = session['user_id']
        
        # Get form data - only save general settings (Azure credentials are in .env)
        settings = {
            'ad_enabled': '1' if request.form.get('enabled') == 'on' else '0',
            'ad_auth_method': 'oauth',  # Always OAuth for Azure AD
            
            # Common Settings
            'ad_admin_group': request.form.get('admin_group', '').strip(),
            'ad_manager_group': request.form.get('manager_group', '').strip(),
            'ad_auto_create_users': '1' if request.form.get('auto_create_users') == 'on' else '0',
            'ad_default_hativa_id': request.form.get('default_hativa_id', '').strip(),
            'ad_sync_on_login': '1' if request.form.get('sync_on_login') == 'on' else '0'
        }
        
        # Update all settings
        for key, value in settings.items():
            db.update_system_setting(key, value, user_id)
        
        # Reload AD service with new settings
        ad_service.reload_settings()
        
        audit_logger.log_success(
            audit_logger.ACTION_UPDATE,
            'ad_settings',
            details='עדכון הגדרות Active Directory'
        )
        
        flash('הגדרות Active Directory עודכנו בהצלחה', 'success')
        
    except Exception as e:
        audit_logger.log_error(
            audit_logger.ACTION_UPDATE,
            'ad_settings',
            str(e),
            details='עדכון הגדרות Active Directory'
        )
        flash(f'שגיאה בעדכון הגדרות AD: {str(e)}', 'error')
    
    return redirect(url_for('settings.ad_settings'))

@settings_bp.route('/admin/ad_settings/test', methods=['POST'])
@admin_required
def test_ad_connection():
    """Test Active Directory connection"""
    try:
        # Reload settings first
        ad_service.reload_settings()
        
        # Test connection
        success, message = ad_service.test_connection()
        
        if success:
            audit_logger.log_success(
                audit_logger.ACTION_TEST,
                'ad_connection',
                details='בדיקת חיבור AD הצליחה'
            )
            return jsonify({'success': True, 'message': message})
        else:
            audit_logger.log_error(
                audit_logger.ACTION_TEST,
                'ad_connection',
                message,
                details='בדיקת חיבור AD נכשלה'
            )
            return jsonify({'success': False, 'message': message})
            
    except Exception as e:
        current_app.logger.error(f'Error testing AD connection: {str(e)}')
        return jsonify({'success': False, 'message': f'שגיאה: {str(e)}'})

@settings_bp.route('/admin/ad_settings/test_azure', methods=['POST'])
@admin_required
def test_azure_connection():
    """Test Azure AD OAuth configuration"""
    try:
        # Reload settings first
        ad_service.reload_settings()
        
        # Test Azure AD configuration
        success, message = ad_service.test_azure_connection()
        
        if success:
            audit_logger.log_success(
                audit_logger.ACTION_TEST,
                'azure_ad_config',
                details='בדיקת הגדרות Azure AD הצליחה'
            )
            return jsonify({'success': True, 'message': message})
        else:
            audit_logger.log_error(
                audit_logger.ACTION_TEST,
                'azure_ad_config',
                message,
                details='בדיקת הגדרות Azure AD נכשלה'
            )
            return jsonify({'success': False, 'message': message})
            
    except Exception as e:
        current_app.logger.error(f'Error testing Azure AD configuration: {str(e)}')
        return jsonify({'success': False, 'message': f'שגיאה: {str(e)}'})

@settings_bp.route('/admin/ad_settings/search_users', methods=['POST'])
@admin_required
def search_ad_users():
    """Search for users in Active Directory"""
    try:
        search_term = request.form.get('search_term', '').strip()
        
        if not search_term:
            return jsonify({'success': False, 'message': 'נדרש מונח חיפוש'})
        
        # Search AD
        users = ad_service.search_users(search_term, limit=20)
        
        return jsonify({'success': True, 'users': users, 'count': len(users)})
        
    except Exception as e:
        current_app.logger.error(f'Error searching AD users: {str(e)}')
        return jsonify({'success': False, 'message': f'שגיאה בחיפוש: {str(e)}'})

@settings_bp.route('/admin/ad_settings/sync_user', methods=['POST'])
@admin_required
def sync_ad_user():
    """Manually sync a user from AD"""
    try:
        username = request.form.get('username', '').strip()
        role = request.form.get('role', 'user')
        hativa_id = request.form.get('hativa_id')
        
        if not username:
            return jsonify({'success': False, 'message': 'נדרש שם משתמש'})
        
        # Get user from AD
        _, user_info, _ = ad_service.authenticate(username, '')
        
        if not user_info:
            # Try to search for user
            users = ad_service.search_users(username, limit=1)
            if users:
                user_info = users[0]
            else:
                return jsonify({'success': False, 'message': 'משתמש לא נמצא ב-AD'})
        
        # Sync to local DB
        hativa_id_int = int(hativa_id) if hativa_id else None
        user_id = ad_service.sync_user_to_local(user_info, role, hativa_id_int)
        
        if user_id:
            audit_logger.log_user_created(user_id, username, role)
            return jsonify({'success': True, 'message': f'משתמש {username} סונכרן בהצלחה'})
        else:
            return jsonify({'success': False, 'message': 'שגיאה בסנכרון המשתמש'})
            
    except Exception as e:
        current_app.logger.error(f'Error syncing AD user: {str(e)}')
        return jsonify({'success': False, 'message': f'שגיאה: {str(e)}'})
